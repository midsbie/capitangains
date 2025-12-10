import datetime as dt
from decimal import Decimal

from openpyxl import load_workbook

from capitangains.reporting.fifo import RealizedLine
from capitangains.reporting.fifo_domain import SellMatchLeg
from capitangains.reporting.fx import FxTable
from capitangains.reporting.report_builder import ReportBuilder
from capitangains.reporting.report_sink import ExcelReportSink


def make_fx() -> FxTable:
    ft = FxTable()
    # EUR identity is handled internally; add USD for a couple dates
    ft.data["USD"]["2024-01-10"] = Decimal("0.9")
    ft.data["USD"]["2024-01-05"] = Decimal("0.9")
    ft.date_index["USD"] = sorted(ft.data["USD"].keys())
    return ft


def build_rb_for_summary():
    rb = ReportBuilder(year=2024)
    # One EUR trade
    rl_eur = RealizedLine(
        symbol="ASML",
        currency="EUR",
        sell_date=dt.date(2024, 1, 5),
        sell_qty=Decimal("10"),
        sell_gross_ccy=Decimal("1000"),
        sell_comm_ccy=Decimal("-1.00"),
        sell_net_ccy=Decimal("999.00"),
        legs=[
            SellMatchLeg(
                buy_date=dt.date(2023, 12, 1),
                qty=Decimal("10"),
                lot_qty_before=Decimal("10"),
                alloc_cost_ccy=Decimal("800.00"),
            )
        ],
        realized_pl_ccy=Decimal("199.00"),
    )
    rb.add_realized(rl_eur)

    # One USD trade
    rl_usd = RealizedLine(
        symbol="AMD",
        currency="USD",
        sell_date=dt.date(2024, 1, 10),
        sell_qty=Decimal("5"),
        sell_gross_ccy=Decimal("500"),
        sell_comm_ccy=Decimal("0"),
        sell_net_ccy=Decimal("500"),
        legs=[
            SellMatchLeg(
                buy_date=dt.date(2023, 6, 1),
                qty=Decimal("5"),
                lot_qty_before=Decimal("5"),
                alloc_cost_ccy=Decimal("400.00"),
            )
        ],
        realized_pl_ccy=Decimal("100.00"),
    )
    rb.add_realized(rl_usd)

    # Convert to EUR
    rb.convert_eur(make_fx())
    return rb


def test_summary_sheet_contents(tmp_path):
    rb = build_rb_for_summary()
    out = tmp_path / "out.xlsx"
    sink = ExcelReportSink(out_path=out, locale="EN")
    sink.write(rb)
    wb = load_workbook(out)
    ws = wb["Trading Totals"]
    rows = list(ws.iter_rows(values_only=True))
    # Expect header + at least 4 lines: total P/L EUR, proceeds EUR, alloc EUR, USD breakdown
    labels = [r[0] for r in rows[1:4]]
    assert "Total Realized P/L (EUR)" in labels
    assert "Total Net Proceeds (EUR)" in labels
    assert "Total Allocated Cost (EUR)" in labels
    # No duplicate EUR total lines
    assert labels.count("Total Realized P/L (EUR)") == 1

    # Non-EUR breakdown includes USD but not EUR
    breakdown_labels = [r[0] for r in rows[4:]]
    assert any("Total Realized P/L (USD)" == s for s in breakdown_labels)
    assert not any("Total Realized P/L (EUR)" == s for s in breakdown_labels)


def test_per_symbol_number_formats(tmp_path):
    rb = build_rb_for_summary()
    out = tmp_path / "out.xlsx"
    sink = ExcelReportSink(out_path=out, locale="EN")
    sink.write(rb)
    wb = load_workbook(out)
    ws = wb["Per Symbol Summary"]
    # Find AMD (USD) row and assert number formats
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "AMD":
            # Trade currency columns: 3,4,5 (0-based indices 2..4) should be USD
            assert row[2].number_format.startswith("$")
            assert row[3].number_format.startswith("$")
            assert row[4].number_format.startswith("$")
            # EUR columns 6,7,8 should be EUR format
            assert row[5].number_format.startswith("€")
            assert row[6].number_format.startswith("€")
            assert row[7].number_format.startswith("€")
            break
    else:
        assert False, "AMD row not found"

    # Find ASML (EUR) row and assert formats are EUR across both sets
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "ASML":
            assert row[2].number_format.startswith("€")
            assert row[3].number_format.startswith("€")
            assert row[4].number_format.startswith("€")
            assert row[5].number_format.startswith("€")
            assert row[6].number_format.startswith("€")
            assert row[7].number_format.startswith("€")
            break
    else:
        assert False, "ASML row not found"
