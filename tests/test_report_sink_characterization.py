"""Per-cell characterization of the eight uniform table sheets in ExcelReportSink.

The sink was refactored from eight near-identical _write_* methods into one data-driven
engine plus a declarative spec per sheet. End-to-end equivalence rests on producing an
identical workbook, and the repo has no golden-file infrastructure, so the safety net is
this explicit (value, number_format) matrix: for every header and data cell of every
table sheet, under both locales, both the loaded value and the applied number format are
pinned against literals captured from the pre-refactor output.

Cells whose expected format is "General" are the TEXT cells; pinning them locks the
load-bearing guarantee that the engine never assigns a number format to a text cell
(openpyxl's default "General" must survive untouched, a byte-preservation detail).

The summary sheet is intentionally excluded: it is a metric/value table left as-is by
the refactor, and its labels and values are covered by the test_summary_and_formats
module (its money number formats reuse the same _NumberFormats.money exercised
here). Its presence and position are still locked via _SHEET_ORDER.
"""

import datetime as dt
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from test_report_builder_sink import _make_fx, _realized

from capitangains.reporting.extract import (
    DividendRow,
    InterestRow,
    SyepInterestRow,
    TransferRow,
    WithholdingRow,
)
from capitangains.reporting.report_builder import ReportBuilder
from capitangains.reporting.report_sink import ExcelReportSink, _NumberFormats


def _build_report() -> ReportBuilder:
    """Build one report populating all eight table sheets with the edge cases the
    refactor must preserve.

    Exercised: every per-sheet sort key (in deliberately non-sorted insertion order); a
    multi-leg realized line (two anexo_j rows + a proceeds split) with one transferred
    and one synthetic leg; a gap line (gap_status + per_symbol has_gap) whose single leg
    has a None buy date (a nullable date in anexo_j); a line left unconverted because
    its currency has no FX rate (nullable EUR cells in realized + anexo_j); nullable
    cash-flow EUR amounts (GBP dividend/withholding with no rate); and a SYEP row with
    null value/start dates (nullable dates + nullable interest_paid_eur).
    """
    rb = ReportBuilder(year=2024)

    # Realized lines in non-sorted insertion order so realized/anexo_j (insertion order)
    # are distinguishable from per_symbol (sorted by symbol).
    mlt = _realized(
        "MLT",
        "USD",
        dt.date(2024, 6, 15),
        [
            {
                "buy_date": dt.date(2023, 1, 1),
                "qty": Decimal("7"),
                "alloc_cost_ccy": Decimal("40"),
            },
            {
                "buy_date": dt.date(2023, 2, 1),
                "qty": Decimal("3"),
                "alloc_cost_ccy": Decimal("30"),
            },
        ],
    )
    mlt.legs[0].transferred = True
    mlt.legs[1].synthetic = True
    rb.add_realized(mlt)

    rb.add_realized(
        _realized(
            "GAP",
            "USD",
            dt.date(2024, 6, 20),
            [{"buy_date": None, "qty": Decimal("5"), "alloc_cost_ccy": Decimal("0")}],
            has_gap=True,
        )
    )
    rb.add_realized(
        _realized(
            "NOC",
            "GBP",
            dt.date(2024, 6, 30),
            [
                {
                    "buy_date": dt.date(2023, 5, 1),
                    "qty": Decimal("2"),
                    "alloc_cost_ccy": Decimal("20"),
                }
            ],
        )
    )

    rb.set_dividends(
        [
            DividendRow("USD", dt.date(2024, 2, 1), "Zulu", Decimal("5")),
            DividendRow("USD", dt.date(2024, 2, 2), "alpha", Decimal("3")),
            DividendRow("GBP", dt.date(2024, 2, 3), "Beta", Decimal("7")),
        ]
    )
    rb.set_interest(
        [
            InterestRow("USD", dt.date(2024, 3, 1), "Zeta", Decimal("2")),
            InterestRow("USD", dt.date(2024, 3, 2), "apex", Decimal("1")),
        ]
    )
    rb.set_withholding(
        [
            WithholdingRow(
                "EUR", dt.date(2024, 4, 2), "Zulu", Decimal("-1"), "", "Interest", ""
            ),
            WithholdingRow(
                "EUR", dt.date(2024, 4, 3), "Alpha", Decimal("-1.5"), "", "Dividend", ""
            ),
            WithholdingRow(
                "USD", dt.date(2024, 4, 1), "Bravo", Decimal("-2"), "", "Dividend", "US"
            ),
            WithholdingRow(
                "GBP", dt.date(2024, 4, 4), "Gamma", Decimal("-3"), "", "Unknown", "GB"
            ),
        ]
    )
    rb.set_syep_interest(
        [
            SyepInterestRow(
                currency="USD",
                value_date=dt.date(2024, 5, 1),
                symbol="LEND",
                start_date=dt.date(2023, 12, 1),
                quantity=Decimal("-10"),
                collateral_amount=Decimal("1000"),
                market_rate_pct=Decimal("5.25"),
                customer_rate_pct=Decimal("2.5"),
                interest_paid=Decimal("3"),
                code="",
            ),
            SyepInterestRow(
                currency="USD",
                value_date=None,
                symbol="NODT",
                start_date=None,
                quantity=Decimal("-5"),
                collateral_amount=Decimal("500"),
                market_rate_pct=Decimal("4"),
                customer_rate_pct=Decimal("1.5"),
                interest_paid=Decimal("1"),
                code="Po",
            ),
        ]
    )
    rb.set_transfers(
        [
            TransferRow(
                section="Transfers",
                asset_category="Stocks",
                currency="USD",
                symbol="ZZZ",
                date=dt.date(2024, 3, 1),
                direction="Out",
                quantity=Decimal("5"),
                market_value=Decimal("0"),
                code="",
            ),
            TransferRow(
                section="Transfers",
                asset_category="Stocks",
                currency="USD",
                symbol="AAA",
                date=dt.date(2024, 1, 1),
                direction="In",
                quantity=Decimal("10"),
                market_value=Decimal("1000"),
                code="FX",
            ),
            TransferRow(
                section="Transfers",
                asset_category="Stocks",
                currency="EUR",
                symbol="BBB",
                date=dt.date(2024, 1, 1),
                direction="In",
                quantity=Decimal("20"),
                market_value=Decimal("2000"),
                code="",
            ),
        ]
    )

    # USD rates for every USD date in play; GBP deliberately absent (leaves NOC and the
    # GBP cash-flow rows unconverted); EUR is identity inside the converter.
    rb.convert_eur(
        _make_fx(
            {
                ("USD", "2023-01-01"): Decimal("0.9"),
                ("USD", "2023-02-01"): Decimal("0.9"),
                ("USD", "2024-02-01"): Decimal("0.9"),
                ("USD", "2024-02-02"): Decimal("0.9"),
                ("USD", "2024-03-01"): Decimal("0.9"),
                ("USD", "2024-03-02"): Decimal("0.9"),
                ("USD", "2024-04-01"): Decimal("0.9"),
                ("USD", "2024-05-01"): Decimal("0.9"),
                ("USD", "2024-06-15"): Decimal("0.9"),
                ("USD", "2024-06-20"): Decimal("0.9"),
            }
        )
    )
    return rb


# Captured from the pre-refactor sink output; see module docstring. Non-ASCII label and
# currency-symbol text is escaped so the matrix stays byte-exact and ASCII-only.
_SHEET_ORDER: dict[str, list[str]] = {
    "EN": [
        "Trading Totals",
        "Realized Trades",
        "Lot-Level EUR Breakdown",
        "Per Symbol Summary",
        "Dividends",
        "Account Interest",
        "SYEP Interest",
        "Withholding Tax",
        "Stock Transfers",
    ],
    "PT": [
        "Totais de Opera\xe7\xf5es",
        "Opera\xe7\xf5es Realizadas",
        "Opera\xe7\xf5es por Lote (Anexo J)",
        "Resumo por S\xedmbolo",
        "Dividendos",
        "Juros da Conta",
        "Juros SYEP",
        "Reten\xe7\xe3o na Fonte",
        "Transfer\xeancias de A\xe7\xf5es",
    ],
}

_EXPECTED: dict[str, dict[str, list]] = {
    "EN": {
        "Realized Trades": [
            [
                ("Ticker", "General"),
                ("Trade Currency", "General"),
                ("Sell Date", "General"),
                ("Quantity Sold", "General"),
                ("Gross Proceeds (Trade Currency)", "General"),
                ("Commissions/Fees (Trade Currency)", "General"),
                ("Net Proceeds (Trade Currency)", "General"),
                ("Allocated Cost Basis (Trade Currency)", "General"),
                ("Realized P/L (Trade Currency)", "General"),
                ("Gross Proceeds (EUR)", "General"),
                ("Commissions/Fees (EUR)", "General"),
                ("Net Proceeds (EUR)", "General"),
                ("Allocated Cost Basis (EUR)", "General"),
                ("Realized P/L (EUR)", "General"),
                ("Matched Buy Lots (JSON)", "General"),
                ("Basis Status", "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (dt.datetime(2024, 6, 15, 0, 0), "YYYY-MM-DD"),
                (10, "0.########"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (100, "$#,##0.00"),
                (70, "$#,##0.00"),
                (30, "$#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (0, "\u20ac#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (63, "\u20ac#,##0.00"),
                (27, "\u20ac#,##0.00"),
                (
                    '[{"buy_date": "2023-01-01", "qty": "7", "alloc_cost_ccy": "40"}, '
                    '{"buy_date": "2023-02-01", "qty": "3", "alloc_cost_ccy": "30"}]',
                    "General",
                ),
                (None, "General"),
            ],
            [
                ("GAP", "General"),
                ("USD", "General"),
                (dt.datetime(2024, 6, 20, 0, 0), "YYYY-MM-DD"),
                (5, "0.########"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (100, "$#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (0, "\u20ac#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (0, "\u20ac#,##0.00"),
                (90, "\u20ac#,##0.00"),
                ('[{"buy_date": null, "qty": "5", "alloc_cost_ccy": "0"}]', "General"),
                ("zero-cost gap", "General"),
            ],
            [
                ("NOC", "General"),
                ("GBP", "General"),
                (dt.datetime(2024, 6, 30, 0, 0), "YYYY-MM-DD"),
                (2, "0.########"),
                (100, "\xa3#,##0.00"),
                (0, "\xa3#,##0.00"),
                (100, "\xa3#,##0.00"),
                (20, "\xa3#,##0.00"),
                (80, "\xa3#,##0.00"),
                (None, "\u20ac#,##0.00"),
                (None, "\u20ac#,##0.00"),
                (None, "\u20ac#,##0.00"),
                (None, "\u20ac#,##0.00"),
                (None, "\u20ac#,##0.00"),
                (
                    '[{"buy_date": "2023-05-01", "qty": "2", "alloc_cost_ccy": "20"}]',
                    "General",
                ),
                (None, "General"),
            ],
        ],
        "Lot-Level EUR Breakdown": [
            [
                ("Ticker", "General"),
                ("Trade Currency", "General"),
                ("Acquisition Date", "General"),
                ("Disposal Date", "General"),
                ("Quantity", "General"),
                ("Acquisition Value (EUR)", "General"),
                ("Disposal Value (EUR)", "General"),
                ("Realized P/L (EUR)", "General"),
                ("Transferred", "General"),
                ("Synthetic", "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (dt.datetime(2023, 1, 1, 0, 0), "YYYY-MM-DD"),
                (dt.datetime(2024, 6, 15, 0, 0), "YYYY-MM-DD"),
                (7, "0.########"),
                (36, "\u20ac#,##0.00"),
                (63, "\u20ac#,##0.00"),
                (27, "\u20ac#,##0.00"),
                ("Yes", "General"),
                (None, "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (dt.datetime(2023, 2, 1, 0, 0), "YYYY-MM-DD"),
                (dt.datetime(2024, 6, 15, 0, 0), "YYYY-MM-DD"),
                (3, "0.########"),
                (27, "\u20ac#,##0.00"),
                (27, "\u20ac#,##0.00"),
                (0, "\u20ac#,##0.00"),
                (None, "General"),
                ("Yes", "General"),
            ],
            [
                ("GAP", "General"),
                ("USD", "General"),
                (None, "YYYY-MM-DD"),
                (dt.datetime(2024, 6, 20, 0, 0), "YYYY-MM-DD"),
                (5, "0.########"),
                (0, "\u20ac#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (None, "General"),
                (None, "General"),
            ],
            [
                ("NOC", "General"),
                ("GBP", "General"),
                (dt.datetime(2023, 5, 1, 0, 0), "YYYY-MM-DD"),
                (dt.datetime(2024, 6, 30, 0, 0), "YYYY-MM-DD"),
                (2, "0.########"),
                (None, "\u20ac#,##0.00"),
                (None, "\u20ac#,##0.00"),
                (None, "\u20ac#,##0.00"),
                (None, "General"),
                (None, "General"),
            ],
        ],
        "Per Symbol Summary": [
            [
                ("Ticker", "General"),
                ("Trade Currency", "General"),
                ("Realized P/L (Trade Currency)", "General"),
                ("Net Proceeds (Trade Currency)", "General"),
                ("Allocated Cost Basis (Trade Currency)", "General"),
                ("Realized P/L (EUR)", "General"),
                ("Net Proceeds (EUR)", "General"),
                ("Allocated Cost Basis (EUR)", "General"),
                ("Gap / Synthetic", "General"),
            ],
            [
                ("GAP", "General"),
                ("USD", "General"),
                (100, "$#,##0.00"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (0, "\u20ac#,##0.00"),
                ("Yes", "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (30, "$#,##0.00"),
                (100, "$#,##0.00"),
                (70, "$#,##0.00"),
                (27, "\u20ac#,##0.00"),
                (90, "\u20ac#,##0.00"),
                (63, "\u20ac#,##0.00"),
                (None, "General"),
            ],
            [
                ("NOC", "General"),
                ("GBP", "General"),
                (80, "\xa3#,##0.00"),
                (100, "\xa3#,##0.00"),
                (20, "\xa3#,##0.00"),
                (0, "\u20ac#,##0.00"),
                (0, "\u20ac#,##0.00"),
                (0, "\u20ac#,##0.00"),
                (None, "General"),
            ],
        ],
        "Dividends": [
            [
                ("Date", "General"),
                ("Currency", "General"),
                ("Description", "General"),
                ("Amount (Currency)", "General"),
                ("Amount (EUR)", "General"),
            ],
            [
                (dt.datetime(2024, 2, 2, 0, 0), "YYYY-MM-DD"),
                ("USD", "General"),
                ("alpha", "General"),
                (3, "$#,##0.00"),
                (2.7, "\u20ac#,##0.00"),
            ],
            [
                (dt.datetime(2024, 2, 3, 0, 0), "YYYY-MM-DD"),
                ("GBP", "General"),
                ("Beta", "General"),
                (7, "\xa3#,##0.00"),
                (None, "\u20ac#,##0.00"),
            ],
            [
                (dt.datetime(2024, 2, 1, 0, 0), "YYYY-MM-DD"),
                ("USD", "General"),
                ("Zulu", "General"),
                (5, "$#,##0.00"),
                (4.5, "\u20ac#,##0.00"),
            ],
        ],
        "Account Interest": [
            [
                ("Date", "General"),
                ("Currency", "General"),
                ("Description", "General"),
                ("Amount (Currency)", "General"),
                ("Amount (EUR)", "General"),
            ],
            [
                (dt.datetime(2024, 3, 2, 0, 0), "YYYY-MM-DD"),
                ("USD", "General"),
                ("apex", "General"),
                (1, "$#,##0.00"),
                (0.9, "\u20ac#,##0.00"),
            ],
            [
                (dt.datetime(2024, 3, 1, 0, 0), "YYYY-MM-DD"),
                ("USD", "General"),
                ("Zeta", "General"),
                (2, "$#,##0.00"),
                (1.8, "\u20ac#,##0.00"),
            ],
        ],
        "SYEP Interest": [
            [
                ("Value Date", "General"),
                ("Currency", "General"),
                ("Symbol", "General"),
                ("Start Date", "General"),
                ("Quantity", "General"),
                ("Collateral Amount", "General"),
                ("Market Rate (%)", "General"),
                ("Customer Rate (%)", "General"),
                ("Interest Paid (Currency)", "General"),
                ("Interest Paid (EUR)", "General"),
                ("Code", "General"),
            ],
            [
                (dt.datetime(2024, 5, 1, 0, 0), "YYYY-MM-DD"),
                ("USD", "General"),
                ("LEND", "General"),
                (dt.datetime(2023, 12, 1, 0, 0), "YYYY-MM-DD"),
                (-10, "0.########"),
                (1000, "$#,##0.00"),
                (5.25, "0.00####"),
                (2.5, "0.00####"),
                (3, "$#,##0.00"),
                (2.7, "\u20ac#,##0.00"),
                (None, "General"),
            ],
            [
                (None, "YYYY-MM-DD"),
                ("USD", "General"),
                ("NODT", "General"),
                (None, "YYYY-MM-DD"),
                (-5, "0.########"),
                (500, "$#,##0.00"),
                (4, "0.00####"),
                (1.5, "0.00####"),
                (1, "$#,##0.00"),
                (None, "\u20ac#,##0.00"),
                ("Po", "General"),
            ],
        ],
        "Withholding Tax": [
            [
                ("Date", "General"),
                ("Currency", "General"),
                ("Description", "General"),
                ("Type", "General"),
                ("Country", "General"),
                ("Amount (Currency)", "General"),
                ("Amount (EUR)", "General"),
            ],
            [
                (dt.datetime(2024, 4, 3, 0, 0), "YYYY-MM-DD"),
                ("EUR", "General"),
                ("Alpha", "General"),
                ("Dividend", "General"),
                (None, "General"),
                (-1.5, "\u20ac#,##0.00"),
                (-1.5, "\u20ac#,##0.00"),
            ],
            [
                (dt.datetime(2024, 4, 2, 0, 0), "YYYY-MM-DD"),
                ("EUR", "General"),
                ("Zulu", "General"),
                ("Interest", "General"),
                (None, "General"),
                (-1, "\u20ac#,##0.00"),
                (-1, "\u20ac#,##0.00"),
            ],
            [
                (dt.datetime(2024, 4, 4, 0, 0), "YYYY-MM-DD"),
                ("GBP", "General"),
                ("Gamma", "General"),
                ("Unknown", "General"),
                ("GB", "General"),
                (-3, "\xa3#,##0.00"),
                (None, "\u20ac#,##0.00"),
            ],
            [
                (dt.datetime(2024, 4, 1, 0, 0), "YYYY-MM-DD"),
                ("USD", "General"),
                ("Bravo", "General"),
                ("Dividend", "General"),
                ("US", "General"),
                (-2, "$#,##0.00"),
                (-1.8, "\u20ac#,##0.00"),
            ],
        ],
        "Stock Transfers": [
            [
                ("Date", "General"),
                ("Symbol", "General"),
                ("Direction", "General"),
                ("Quantity", "General"),
                ("Currency", "General"),
                ("Market Value", "General"),
                ("Code", "General"),
            ],
            [
                (dt.datetime(2024, 1, 1, 0, 0), "YYYY-MM-DD"),
                ("AAA", "General"),
                ("In", "General"),
                (10, "0.########"),
                ("USD", "General"),
                (1000, "$#,##0.00"),
                ("FX", "General"),
            ],
            [
                (dt.datetime(2024, 1, 1, 0, 0), "YYYY-MM-DD"),
                ("BBB", "General"),
                ("In", "General"),
                (20, "0.########"),
                ("EUR", "General"),
                (2000, "\u20ac#,##0.00"),
                (None, "General"),
            ],
            [
                (dt.datetime(2024, 3, 1, 0, 0), "YYYY-MM-DD"),
                ("ZZZ", "General"),
                ("Out", "General"),
                (5, "0.########"),
                ("USD", "General"),
                (0, "$#,##0.00"),
                (None, "General"),
            ],
        ],
    },
    "PT": {
        "Opera\xe7\xf5es Realizadas": [
            [
                ("S\xedmbolo", "General"),
                ("Moeda da Opera\xe7\xe3o", "General"),
                ("Data de Venda", "General"),
                ("Quantidade Vendida", "General"),
                ("Proveitos Brutos (Moeda)", "General"),
                ("Comiss\xf5es/Taxas (Moeda)", "General"),
                ("Proveitos L\xedquidos (Moeda)", "General"),
                ("Custo Alocado (Moeda)", "General"),
                ("Resultado Realizado (Moeda)", "General"),
                ("Proveitos Brutos (EUR)", "General"),
                ("Comiss\xf5es/Taxas (EUR)", "General"),
                ("Proveitos L\xedquidos (EUR)", "General"),
                ("Custo Alocado (EUR)", "General"),
                ("Resultado Realizado (EUR)", "General"),
                ("Lotes de Compra (JSON)", "General"),
                ("Estado do Custo", "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (dt.datetime(2024, 6, 15, 0, 0), "DD/MM/YYYY"),
                (10, "0.########"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (100, "$#,##0.00"),
                (70, "$#,##0.00"),
                (30, "$#,##0.00"),
                (90, '#,##0.00 "\u20ac"'),
                (0, '#,##0.00 "\u20ac"'),
                (90, '#,##0.00 "\u20ac"'),
                (63, '#,##0.00 "\u20ac"'),
                (27, '#,##0.00 "\u20ac"'),
                (
                    '[{"buy_date": "2023-01-01", "qty": "7", "alloc_cost_ccy": "40"}, '
                    '{"buy_date": "2023-02-01", "qty": "3", "alloc_cost_ccy": "30"}]',
                    "General",
                ),
                (None, "General"),
            ],
            [
                ("GAP", "General"),
                ("USD", "General"),
                (dt.datetime(2024, 6, 20, 0, 0), "DD/MM/YYYY"),
                (5, "0.########"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (100, "$#,##0.00"),
                (90, '#,##0.00 "\u20ac"'),
                (0, '#,##0.00 "\u20ac"'),
                (90, '#,##0.00 "\u20ac"'),
                (0, '#,##0.00 "\u20ac"'),
                (90, '#,##0.00 "\u20ac"'),
                ('[{"buy_date": null, "qty": "5", "alloc_cost_ccy": "0"}]', "General"),
                ("zero-cost gap", "General"),
            ],
            [
                ("NOC", "General"),
                ("GBP", "General"),
                (dt.datetime(2024, 6, 30, 0, 0), "DD/MM/YYYY"),
                (2, "0.########"),
                (100, "\xa3#,##0.00"),
                (0, "\xa3#,##0.00"),
                (100, "\xa3#,##0.00"),
                (20, "\xa3#,##0.00"),
                (80, "\xa3#,##0.00"),
                (None, '#,##0.00 "\u20ac"'),
                (None, '#,##0.00 "\u20ac"'),
                (None, '#,##0.00 "\u20ac"'),
                (None, '#,##0.00 "\u20ac"'),
                (None, '#,##0.00 "\u20ac"'),
                (
                    '[{"buy_date": "2023-05-01", "qty": "2", "alloc_cost_ccy": "20"}]',
                    "General",
                ),
                (None, "General"),
            ],
        ],
        "Opera\xe7\xf5es por Lote (Anexo J)": [
            [
                ("S\xedmbolo", "General"),
                ("Moeda da Opera\xe7\xe3o", "General"),
                ("Data de Aquisi\xe7\xe3o", "General"),
                ("Data de Venda", "General"),
                ("Quantidade", "General"),
                ("Valor de Aquisi\xe7\xe3o (EUR)", "General"),
                ("Valor de Realiza\xe7\xe3o (EUR)", "General"),
                ("Mais/menos\u2011valia (EUR)", "General"),
                ("Transferido", "General"),
                ("Sint\xe9tico", "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (dt.datetime(2023, 1, 1, 0, 0), "DD/MM/YYYY"),
                (dt.datetime(2024, 6, 15, 0, 0), "DD/MM/YYYY"),
                (7, "0.########"),
                (36, '#,##0.00 "\u20ac"'),
                (63, '#,##0.00 "\u20ac"'),
                (27, '#,##0.00 "\u20ac"'),
                ("Yes", "General"),
                (None, "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (dt.datetime(2023, 2, 1, 0, 0), "DD/MM/YYYY"),
                (dt.datetime(2024, 6, 15, 0, 0), "DD/MM/YYYY"),
                (3, "0.########"),
                (27, '#,##0.00 "\u20ac"'),
                (27, '#,##0.00 "\u20ac"'),
                (0, '#,##0.00 "\u20ac"'),
                (None, "General"),
                ("Yes", "General"),
            ],
            [
                ("GAP", "General"),
                ("USD", "General"),
                (None, "DD/MM/YYYY"),
                (dt.datetime(2024, 6, 20, 0, 0), "DD/MM/YYYY"),
                (5, "0.########"),
                (0, '#,##0.00 "\u20ac"'),
                (90, '#,##0.00 "\u20ac"'),
                (90, '#,##0.00 "\u20ac"'),
                (None, "General"),
                (None, "General"),
            ],
            [
                ("NOC", "General"),
                ("GBP", "General"),
                (dt.datetime(2023, 5, 1, 0, 0), "DD/MM/YYYY"),
                (dt.datetime(2024, 6, 30, 0, 0), "DD/MM/YYYY"),
                (2, "0.########"),
                (None, '#,##0.00 "\u20ac"'),
                (None, '#,##0.00 "\u20ac"'),
                (None, '#,##0.00 "\u20ac"'),
                (None, "General"),
                (None, "General"),
            ],
        ],
        "Resumo por S\xedmbolo": [
            [
                ("S\xedmbolo", "General"),
                ("Moeda da Opera\xe7\xe3o", "General"),
                ("Resultado Realizado (Moeda)", "General"),
                ("Proveitos L\xedquidos (Moeda)", "General"),
                ("Custo Alocado (Moeda)", "General"),
                ("Resultado Realizado (EUR)", "General"),
                ("Proveitos L\xedquidos (EUR)", "General"),
                ("Custo Alocado (EUR)", "General"),
                ("Lacuna / Sint\xe9tico", "General"),
            ],
            [
                ("GAP", "General"),
                ("USD", "General"),
                (100, "$#,##0.00"),
                (100, "$#,##0.00"),
                (0, "$#,##0.00"),
                (90, '#,##0.00 "\u20ac"'),
                (90, '#,##0.00 "\u20ac"'),
                (0, '#,##0.00 "\u20ac"'),
                ("Yes", "General"),
            ],
            [
                ("MLT", "General"),
                ("USD", "General"),
                (30, "$#,##0.00"),
                (100, "$#,##0.00"),
                (70, "$#,##0.00"),
                (27, '#,##0.00 "\u20ac"'),
                (90, '#,##0.00 "\u20ac"'),
                (63, '#,##0.00 "\u20ac"'),
                (None, "General"),
            ],
            [
                ("NOC", "General"),
                ("GBP", "General"),
                (80, "\xa3#,##0.00"),
                (100, "\xa3#,##0.00"),
                (20, "\xa3#,##0.00"),
                (0, '#,##0.00 "\u20ac"'),
                (0, '#,##0.00 "\u20ac"'),
                (0, '#,##0.00 "\u20ac"'),
                (None, "General"),
            ],
        ],
        "Dividendos": [
            [
                ("Data", "General"),
                ("Moeda", "General"),
                ("Descri\xe7\xe3o", "General"),
                ("Montante (Moeda)", "General"),
                ("Montante (EUR)", "General"),
            ],
            [
                (dt.datetime(2024, 2, 2, 0, 0), "DD/MM/YYYY"),
                ("USD", "General"),
                ("alpha", "General"),
                (3, "$#,##0.00"),
                (2.7, '#,##0.00 "\u20ac"'),
            ],
            [
                (dt.datetime(2024, 2, 3, 0, 0), "DD/MM/YYYY"),
                ("GBP", "General"),
                ("Beta", "General"),
                (7, "\xa3#,##0.00"),
                (None, '#,##0.00 "\u20ac"'),
            ],
            [
                (dt.datetime(2024, 2, 1, 0, 0), "DD/MM/YYYY"),
                ("USD", "General"),
                ("Zulu", "General"),
                (5, "$#,##0.00"),
                (4.5, '#,##0.00 "\u20ac"'),
            ],
        ],
        "Juros da Conta": [
            [
                ("Data", "General"),
                ("Moeda", "General"),
                ("Descri\xe7\xe3o", "General"),
                ("Montante (Moeda)", "General"),
                ("Montante (EUR)", "General"),
            ],
            [
                (dt.datetime(2024, 3, 2, 0, 0), "DD/MM/YYYY"),
                ("USD", "General"),
                ("apex", "General"),
                (1, "$#,##0.00"),
                (0.9, '#,##0.00 "\u20ac"'),
            ],
            [
                (dt.datetime(2024, 3, 1, 0, 0), "DD/MM/YYYY"),
                ("USD", "General"),
                ("Zeta", "General"),
                (2, "$#,##0.00"),
                (1.8, '#,##0.00 "\u20ac"'),
            ],
        ],
        "Juros SYEP": [
            [
                ("Data", "General"),
                ("Moeda", "General"),
                ("S\xedmbolo", "General"),
                ("Data de In\xedcio", "General"),
                ("Quantidade", "General"),
                ("Valor de Colateral", "General"),
                ("Taxa de Mercado (%)", "General"),
                ("Taxa ao Cliente (%)", "General"),
                ("Juros Pagos (Moeda)", "General"),
                ("Juros Pagos (EUR)", "General"),
                ("C\xf3digo", "General"),
            ],
            [
                (dt.datetime(2024, 5, 1, 0, 0), "DD/MM/YYYY"),
                ("USD", "General"),
                ("LEND", "General"),
                (dt.datetime(2023, 12, 1, 0, 0), "DD/MM/YYYY"),
                (-10, "0.########"),
                (1000, "$#,##0.00"),
                (5.25, "0.00####"),
                (2.5, "0.00####"),
                (3, "$#,##0.00"),
                (2.7, '#,##0.00 "\u20ac"'),
                (None, "General"),
            ],
            [
                (None, "DD/MM/YYYY"),
                ("USD", "General"),
                ("NODT", "General"),
                (None, "DD/MM/YYYY"),
                (-5, "0.########"),
                (500, "$#,##0.00"),
                (4, "0.00####"),
                (1.5, "0.00####"),
                (1, "$#,##0.00"),
                (None, '#,##0.00 "\u20ac"'),
                ("Po", "General"),
            ],
        ],
        "Reten\xe7\xe3o na Fonte": [
            [
                ("Data", "General"),
                ("Moeda", "General"),
                ("Descri\xe7\xe3o", "General"),
                ("Tipo", "General"),
                ("Pa\xeds", "General"),
                ("Montante (Moeda)", "General"),
                ("Montante (EUR)", "General"),
            ],
            [
                (dt.datetime(2024, 4, 3, 0, 0), "DD/MM/YYYY"),
                ("EUR", "General"),
                ("Alpha", "General"),
                ("Dividend", "General"),
                (None, "General"),
                (-1.5, '#,##0.00 "\u20ac"'),
                (-1.5, '#,##0.00 "\u20ac"'),
            ],
            [
                (dt.datetime(2024, 4, 2, 0, 0), "DD/MM/YYYY"),
                ("EUR", "General"),
                ("Zulu", "General"),
                ("Interest", "General"),
                (None, "General"),
                (-1, '#,##0.00 "\u20ac"'),
                (-1, '#,##0.00 "\u20ac"'),
            ],
            [
                (dt.datetime(2024, 4, 4, 0, 0), "DD/MM/YYYY"),
                ("GBP", "General"),
                ("Gamma", "General"),
                ("Unknown", "General"),
                ("GB", "General"),
                (-3, "\xa3#,##0.00"),
                (None, '#,##0.00 "\u20ac"'),
            ],
            [
                (dt.datetime(2024, 4, 1, 0, 0), "DD/MM/YYYY"),
                ("USD", "General"),
                ("Bravo", "General"),
                ("Dividend", "General"),
                ("US", "General"),
                (-2, "$#,##0.00"),
                (-1.8, '#,##0.00 "\u20ac"'),
            ],
        ],
        "Transfer\xeancias de A\xe7\xf5es": [
            [
                ("Data", "General"),
                ("S\xedmbolo", "General"),
                ("Dire\xe7\xe3o", "General"),
                ("Quantidade", "General"),
                ("Moeda", "General"),
                ("Valor de Mercado", "General"),
                ("C\xf3digo", "General"),
            ],
            [
                (dt.datetime(2024, 1, 1, 0, 0), "DD/MM/YYYY"),
                ("AAA", "General"),
                ("In", "General"),
                (10, "0.########"),
                ("USD", "General"),
                (1000, "$#,##0.00"),
                ("FX", "General"),
            ],
            [
                (dt.datetime(2024, 1, 1, 0, 0), "DD/MM/YYYY"),
                ("BBB", "General"),
                ("In", "General"),
                (20, "0.########"),
                ("EUR", "General"),
                (2000, '#,##0.00 "\u20ac"'),
                (None, "General"),
            ],
            [
                (dt.datetime(2024, 3, 1, 0, 0), "DD/MM/YYYY"),
                ("ZZZ", "General"),
                ("Out", "General"),
                (5, "0.########"),
                ("USD", "General"),
                (0, "$#,##0.00"),
                (None, "General"),
            ],
        ],
    },
}


def _grid(ws) -> list:
    return [
        [(cell.value, cell.number_format) for cell in row] for row in ws.iter_rows()
    ]


@pytest.mark.parametrize("locale", ["EN", "PT"])
def test_table_sheets_characterization(tmp_path, locale):
    rb = _build_report()
    out = tmp_path / f"report_{locale}.xlsx"
    ExcelReportSink(out_path=out, locale=locale).write(rb)
    wb = load_workbook(out)

    assert wb.sheetnames == _SHEET_ORDER[locale]

    for title, exp_rows in _EXPECTED[locale].items():
        actual = _grid(wb[title])
        assert len(actual) == len(exp_rows), f"[{locale}] {title}: row count"
        for r, (exp_row, act_row) in enumerate(
            zip(exp_rows, actual, strict=True), start=1
        ):
            assert act_row == exp_row, f"[{locale}] {title} row {r}"


def test_optional_sheets_omitted_when_source_empty(tmp_path):
    """skip_if_empty: with every optional collection empty, only the four always-on
    sheets are written (summary, realized, anexo_j, per_symbol), in order. The five
    optional sheets (dividends, interest, syep, withholding, transfers) are absent.
    """
    rb = ReportBuilder(year=2024)
    rb.add_realized(
        _realized(
            "ONLY",
            "USD",
            dt.date(2024, 6, 1),
            [
                {
                    "buy_date": dt.date(2023, 1, 1),
                    "qty": Decimal("1"),
                    "alloc_cost_ccy": Decimal("10"),
                }
            ],
        )
    )
    out = tmp_path / "sparse.xlsx"
    ExcelReportSink(out_path=out, locale="EN").write(rb)
    assert load_workbook(out).sheetnames == [
        "Trading Totals",
        "Realized Trades",
        "Lot-Level EUR Breakdown",
        "Per Symbol Summary",
    ]


@pytest.mark.parametrize(
    "locale,ccy,expected",
    [
        ("EN", "JPY", "\u00a5#,##0.00"),  # symbol-mapped, non-EUR: prefixed symbol
        ("PT", "JPY", "\u00a5#,##0.00"),  # JPY is not special-cased like EUR/PT
        ("EN", "CHF", '"CHF" #,##0.00'),  # unmapped currency: quoted ISO, EN order
        ("PT", "CHF", '#,##0.00 "CHF"'),  # unmapped currency: quoted ISO, PT order
    ],
)
def test_money_format_fallbacks(locale, ccy, expected):
    assert _NumberFormats(locale).money(ccy) == expected
