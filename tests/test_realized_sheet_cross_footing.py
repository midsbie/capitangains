"""The Realized sheet's money columns must cross-foot (regression guard for F5).

The Realized Trades sheet is the tool's audit surface: a filer checks a figure by
confirming the columns on its row add up. Two additive identities state that contract,
in each of the trade-currency and EUR money blocks:

    gross + fees == net        net - alloc == pl

F5 was that they did not hold: each cell was quantized to cents independently, while the
stored figures had been quantized earlier from raw, sub-cent inputs, so the two
roundings disagreed by a cent on boundary values. The fix makes every block derive its
dependent columns from cent-rounded primitives (allocated cost rounded once on the line;
EUR commission taken as net - gross), so the identities hold by construction. The two
fixtures below are the boundary cases that exposed each half of the defect:

- A leg sum landing on a half-cent (10.001 + 10.004 = 20.005) rounds to 20.01, and the
  P/L derives from that rounded cost, so the trade-currency block foots (50.00 - 20.01 =
  29.99) instead of showing a pl of 30.00 beside an alloc of 20.01.
- A small disposal priced at 0.8765 would round gross, fees, and net apart; deriving the
  EUR fee as net - gross keeps gross + fees == net exactly.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl import load_workbook

from capitangains.reporting.report_builder import ReportBuilder
from capitangains.reporting.report_sink import ExcelReportSink
from tests.support import make_fx, realized_line

# Column offsets of the five trade-currency and five EUR money cells on the Realized
# sheet, in _REALIZED_SPEC order: ticker, currency, sell_date, qty, then the two blocks.
_TCY = slice(4, 9)  # gross, fees, net, alloc, pl  (trade currency)
_EUR = slice(9, 14)  # gross, fees, net, alloc, pl  (EUR)


def _realized_money_row(
    rb: ReportBuilder, tmp_path
) -> tuple[list[Decimal], list[Decimal]]:
    """Write rb to a workbook and read back the single Realized row's money cells.

    Returns (trade_currency_cells, eur_cells), each [gross, fees, net, alloc, pl] as the
    exact cent Decimals the sink rendered (the cells are floats of 2-dp Decimals, so
    Decimal(str(...)) reconstructs them without binary-float noise).
    """
    out = tmp_path / "report.xlsx"
    ExcelReportSink(out_path=out, locale="EN").write(rb)
    ws = load_workbook(out)["Realized Trades"]
    (row,) = list(ws.iter_rows(min_row=2, values_only=True))  # exactly one data row
    return (
        [Decimal(str(c)) for c in row[_TCY]],
        [Decimal(str(c)) for c in row[_EUR]],
    )


def _assert_cross_foots(tcy: list[Decimal], eur: list[Decimal]) -> None:
    """Both money blocks must satisfy gross + fees == net and net - alloc == pl."""
    for label, (gross, fees, net, alloc, pl) in (("tcy", tcy), ("eur", eur)):
        assert gross + fees == net, (
            f"{label}: gross + fees != net ({gross} + {fees} != {net})"
        )
        assert net - alloc == pl, (
            f"{label}: net - alloc != pl ({net} - {alloc} != {pl})"
        )


def test_realized_trade_currency_columns_cross_foot(tmp_path):
    """Two legs whose 1e-8 allocated costs sum to a half-cent (20.005) still foot on the
    trade-currency block: alloc rounds to 20.01 and pl derives from it (50.00 - 20.01 =
    29.99), so net - alloc == pl exactly.
    """
    rb = ReportBuilder(year=2024)
    rb.add_realized(
        realized_line(
            symbol="ACME",
            # Base currency: converts at rate 1, so the EUR block carries the same
            # figures through its own rounding. It need not match tcy cell for cell (it
            # rounds per leg, 10.00 + 10.00 = 20.00, where tcy sums then rounds once,
            # 20.005 -> 20.01); each block only has to foot against itself.
            currency="EUR",
            sell_date="2024-06-15",
            sell_gross_ccy="50.00",
            sell_comm_ccy="0",
            sell_net_ccy="50.00",
            legs=[
                {"buy_date": "2023-01-01", "qty": "6", "alloc_cost_ccy": "10.00100000"},
                {"buy_date": "2023-02-01", "qty": "4", "alloc_cost_ccy": "10.00400000"},
            ],
        )
    )
    rb.convert_eur(None)

    tcy, eur = _realized_money_row(rb, tmp_path)
    _assert_cross_foots(tcy, eur)


def test_realized_eur_columns_cross_foot(tmp_path):
    """A small disposal priced at 0.8765 foots on the EUR block: the fee is the residual
    net - gross, so quantize(0.07 * r) + fees == net even though pricing the commission
    on its own would round the three columns apart.
    """
    rb = ReportBuilder(year=2024)
    rb.add_realized(
        realized_line(
            symbol="ACME",
            currency="USD",
            sell_date="2024-06-20",
            sell_gross_ccy="0.07",
            sell_comm_ccy="-0.03",
            sell_net_ccy="0.04",
            legs=[{"buy_date": "2023-01-01", "qty": "1", "alloc_cost_ccy": "0"}],
        )
    )
    rb.convert_eur(
        make_fx(
            {
                ("USD", "2024-06-20"): Decimal("0.8765"),  # sell-date rate
                ("USD", "2023-01-01"): Decimal("0.8765"),  # leg buy-date rate
            }
        )
    )

    tcy, eur = _realized_money_row(rb, tmp_path)
    _assert_cross_foots(tcy, eur)
