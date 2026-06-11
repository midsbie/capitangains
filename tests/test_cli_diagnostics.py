"""Aggregate diagnostics and full-pipeline behavior.

Exercises capitangains.diagnostics (the boundary report helpers) and
capitangains.pipeline.run (the orchestration) together:

- parse_acknowledged_gaps: parse the itemized SYMBOL@YYYY-MM-DD acknowledgment spec;
  every malformed token is collected and reported as one DataQualityError.
- report_gap_acknowledgments: two-way tie-out of the gaps found against the operator's
  acknowledgments. Any unacknowledged gap, acknowledged-but-defective Basis, or orphan
  acknowledgment is fatal (exit 2); a clean run emits one audit warning per synthesized
  lot, so synthetic cost basis is never silent.
- run: aborts (exit 2, no workbook) on a malformed spec, a failed gap tie-out, or an FX
  table that cannot supply every rate the EUR report needs; exits 1 on an unreadable or
  unparseable FX table (a setup failure, not a data defect). Under --dry-run it runs the
  full preflight and stops before the write, leaving any existing output untouched.
"""

import csv
import datetime as dt
import logging
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from capitangains.cmd.cli import build_argparser
from capitangains.diagnostics import (
    format_reconciliation_sample,
    parse_acknowledged_gaps,
    report_gap_acknowledgments,
    report_invalid_statements,
    report_ordering_collisions,
    report_orphaned_foreign_tax,
    report_reconciliation,
    report_statement_input_conflicts,
    report_symbol_currency_violations,
    report_unattributed_income,
    report_unrecognized_sections,
)
from capitangains.errors import DataQualityError
from capitangains.model import IbkrStatementCsvParser
from capitangains.pipeline import RunOptions, run
from capitangains.reporting import (
    OrderingCollision,
    Quadro8ALine,
    ReconciliationReport,
    StatementInput,
    SymbolReconciliation,
    UnrecognizedSection,
    detect_ordering_collisions,
    detect_orphaned_foreign_tax,
    detect_statement_input_conflicts,
    detect_unattributed_income,
    detect_unrecognized_sections,
    partition_statements_by_metadata,
)
from capitangains.reporting.extract import StatementMetadata, StatementPeriod
from capitangains.reporting.extract.sections import (
    CONSUMED_SECTIONS,
    IGNORED_SECTIONS,
    SEC_DIVIDENDS,
    SEC_INTEREST,
    SEC_SYEP,
    SEC_TRADES,
    SEC_TRANSFERS,
    SEC_WITHHOLDING,
)
from capitangains.reporting.fifo_domain import GapResolution
from capitangains.reporting.quadro_8a import IncomeKind
from tests.support import (
    SYEP_SECTION,
    TRADES_COLUMNS,
    TRANSFERS_COLUMNS,
    WITHHOLDING_COLUMNS,
    Y2023,
    Y2024,
    header_row,
    make_gap_event,
    parse_model,
    section_table,
    statement_meta_rows,
    trade_data,
    trade_row,
    transfer_data,
    transfer_row,
)


def _gap(*, outcome, symbol="AAPL", date=dt.date(2024, 1, 1), message="no buy history"):
    return make_gap_event(
        symbol=symbol,
        date=date,
        remaining_qty="5",
        message=message,
        outcome=outcome,
    )


# --- parse_acknowledged_gaps ---------------------------------------------------------


def test_parse_acknowledged_gaps_none_and_empty_yield_empty_set():
    assert parse_acknowledged_gaps(None) == frozenset()
    assert parse_acknowledged_gaps("") == frozenset()


def test_parse_acknowledged_gaps_skips_empty_tokens():
    # Leading, trailing, and doubled commas (plus surrounding whitespace) are skipped,
    # not errors; the real key survives.
    assert parse_acknowledged_gaps(" ,BABA@2024-01-01,") == frozenset(
        {("BABA", dt.date(2024, 1, 1))}
    )


def test_parse_acknowledged_gaps_dedupes_repeated_keys():
    assert parse_acknowledged_gaps("BABA@2024-01-01,BABA@2024-01-01") == frozenset(
        {("BABA", dt.date(2024, 1, 1))}
    )


def test_parse_acknowledged_gaps_rejects_malformed_tokens():
    # No "@", empty symbol, empty date, and unparseable date are each malformed.
    for bad in ("BABAnoat", "@2024-01-01", "BABA@", "BABA@notadate"):
        with pytest.raises(DataQualityError):
            parse_acknowledged_gaps(bad)


def test_parse_acknowledged_gaps_lists_every_malformed_token():
    # Accumulate, do not fail fast: both bad tokens appear so the spec is fixed in one
    # pass.
    with pytest.raises(DataQualityError) as exc:
        parse_acknowledged_gaps("BABAnoat,VOD@nope")
    msg = str(exc.value)
    assert "BABAnoat" in msg and "VOD@nope" in msg


# --- report_gap_acknowledgments ------------------------------------------------------


def test_report_gap_acknowledgments_no_gaps_no_acks_is_silent(caplog):
    logger = logging.getLogger("gaps_none")
    with caplog.at_level(logging.INFO):
        report_gap_acknowledgments([], frozenset(), logger)
    assert caplog.records == []


def test_report_gap_acknowledgments_unacknowledged_gap_is_fatal(caplog):
    logger = logging.getLogger("gaps_unack")
    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        report_gap_acknowledgments(
            [_gap(outcome=GapResolution.UNACKNOWLEDGED)], frozenset(), logger
        )

    assert exc.value.code == 2
    assert any("Unacknowledged SELL gap" in r.getMessage() for r in caplog.records)


def test_report_gap_acknowledgments_synthesized_gap_warns(caplog):
    logger = logging.getLogger("gaps_synth")
    acknowledged = frozenset({("AAPL", dt.date(2024, 1, 1))})
    with caplog.at_level(logging.WARNING):
        report_gap_acknowledgments(
            [_gap(outcome=GapResolution.SYNTHESIZED)], acknowledged, logger
        )

    warnings = [r for r in caplog.records if r.levelno == logging.WARNING]
    assert len(warnings) == 1  # only the synthesized lot gets an audit-trail warning
    assert "Synthesized residual lot" in warnings[0].getMessage()


def test_report_gap_acknowledgments_accumulates_every_fatal_category(caplog):
    # One run carrying all three fatal conditions must log every one, then a summary,
    # and exit exactly once. No fail-fast that hides later problems from the operator.
    logger = logging.getLogger("gaps_accum")
    unack = _gap(
        outcome=GapResolution.UNACKNOWLEDGED, symbol="UNACK", date=dt.date(2024, 2, 1)
    )
    defective = _gap(
        outcome=GapResolution.DEFECTIVE,
        symbol="CORRUPT",
        date=dt.date(2024, 3, 1),
        message="IBKR Basis (-9999) inconsistent with its Realized P/L",
    )
    acknowledged = frozenset(
        {
            ("CORRUPT", dt.date(2024, 3, 1)),  # matches the defective gap
            ("GHOST", dt.date(2024, 1, 1)),  # orphan: matches no gap this run
        }
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        report_gap_acknowledgments([unack, defective], acknowledged, logger)

    assert exc.value.code == 2
    messages = [r.getMessage() for r in caplog.records]
    assert any("Unacknowledged SELL gap" in m and "UNACK" in m for m in messages)
    assert any("defective IBKR Basis" in m and "CORRUPT" in m for m in messages)
    assert any("Orphan acknowledgment" in m and "GHOST" in m for m in messages)
    assert any("tie-out failed" in m for m in messages)


# --- symbol-currency violations: reporter --------------------------------------------


def test_report_symbol_currency_violations_silent_when_empty(caplog):
    logger = logging.getLogger("symccy_report_none")
    with caplog.at_level(logging.ERROR):
        report_symbol_currency_violations({}, logger)
    assert caplog.records == []


def test_report_symbol_currency_violations_lists_and_exits(caplog):
    # One ERROR per offending symbol, then a summary, then a single exit 2 -- no
    # fail-fast, mirroring the other boundary reporters.
    logger = logging.getLogger("symccy_report_fatal")
    violations = {
        "ABC": frozenset({"USD", "EUR"}),
        "XYZ": frozenset({"GBP", "USD"}),
    }
    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        report_symbol_currency_violations(violations, logger)

    assert exc.value.code == 2
    messages = [r.getMessage() for r in caplog.records]
    assert any("ABC" in m and "EUR" in m and "USD" in m for m in messages)
    assert any("XYZ" in m and "GBP" in m and "USD" in m for m in messages)
    assert any("symbol-currency uniqueness" in m for m in messages)


# --- same-day event ordering: detector ----------------------------------------------


def _trade_row(symbol, currency, date, qty="10"):
    return trade_row(symbol=symbol, currency=currency, date=date, quantity=qty)


def _transfer_row(symbol, currency, date, direction="In"):
    return transfer_row(
        symbol=symbol,
        currency=currency,
        date=date,
        direction=direction,
        quantity="10",
        market_value="1000",
    )


def test_detect_ordering_collisions_silent_without_collision():
    # A transfer on a different day from the trade (same symbol), and a transfer sharing
    # the trade's day but in a different symbol, are both independent for FIFO. The
    # check is scoped to (symbol, currency, date), so neither is a collision.
    trades = [_trade_row("AAPL", "USD", "2024-06-10")]
    transfers = [
        _transfer_row("AAPL", "USD", "2024-03-15"),
        _transfer_row("MSFT", "USD", "2024-06-10"),
    ]
    assert detect_ordering_collisions(trades, transfers) == []


def test_detect_ordering_collisions_two_same_day_transfers():
    # Two transfers of the same symbol on the same day are equally unorderable -- IBKR
    # timestamps neither -- even with no trade that day.
    transfers = [
        _transfer_row("AAPL", "USD", "2024-06-10", "In"),
        _transfer_row("AAPL", "USD", "2024-06-10", "Out"),
    ]
    assert detect_ordering_collisions([], transfers) == [
        OrderingCollision(
            symbol="AAPL",
            currency="USD",
            date=dt.date(2024, 6, 10),
            n_trades=0,
            n_untimed_trades=0,
            n_transfers=2,
        )
    ]


def test_detect_ordering_collisions_lists_every_collision():
    # Two distinct symbol-days collide (a transfer sharing a timed trade's day in each);
    # both are returned, sorted deterministically.
    trades = [
        _trade_row("AAPL", "USD", "2024-06-10"),
        _trade_row("VOD", "GBP", "2024-07-01"),
    ]
    transfers = [
        _transfer_row("AAPL", "USD", "2024-06-10"),
        _transfer_row("VOD", "GBP", "2024-07-01"),
    ]
    collisions = detect_ordering_collisions(trades, transfers)
    assert [(c.symbol, c.n_trades, c.n_transfers) for c in collisions] == [
        ("AAPL", 1, 1),
        ("VOD", 1, 1),
    ]


def test_detect_ordering_collisions_date_only_trade_collides_with_timed_trade():
    # A date-only trade (no intraday time) shares a symbol-day with a timestamped trade.
    # Their FIFO order is not in the data, so it is a collision with no transfer at all.
    trades = [
        _trade_row("AAPL", "USD", "2024-06-15"),  # builder stamps an intraday time
        trade_row(
            symbol="AAPL",
            currency="USD",
            date="2024-06-15",
            datetime_str="2024-06-15",
            quantity="-10",
        ),
    ]
    assert detect_ordering_collisions(trades, []) == [
        OrderingCollision(
            symbol="AAPL",
            currency="USD",
            date=dt.date(2024, 6, 15),
            n_trades=2,
            n_untimed_trades=1,
            n_transfers=0,
        )
    ]


def test_detect_ordering_collisions_two_timed_trades_are_orderable():
    # Two fully timestamped trades on the same symbol-day are ordered by their times, so
    # this is NOT a collision: the detector must not over-abort an ordinary trading day.
    trades = [
        trade_row(
            symbol="AAPL",
            currency="USD",
            date="2024-06-15",
            datetime_str="2024-06-15, 09:30:00",
            quantity="10",
        ),
        trade_row(
            symbol="AAPL",
            currency="USD",
            date="2024-06-15",
            datetime_str="2024-06-15, 15:00:00",
            quantity="-10",
        ),
    ]
    assert detect_ordering_collisions(trades, []) == []


def test_detect_ordering_collisions_lone_date_only_trade_is_silent():
    # A single date-only trade has nothing to be ordered against, so it is orderable.
    trades = [
        trade_row(
            symbol="AAPL",
            currency="USD",
            date="2024-06-15",
            datetime_str="2024-06-15",
            quantity="-10",
        )
    ]
    assert detect_ordering_collisions(trades, []) == []


# --- same-day event ordering: reporter -----------------------------------------------


def test_report_ordering_collisions_silent_when_empty(caplog):
    logger = logging.getLogger("ordering_report_none")
    with caplog.at_level(logging.ERROR):
        report_ordering_collisions([], logger)
    assert caplog.records == []


def test_report_ordering_collisions_lists_and_exits(caplog):
    # Every collision is named, then a single summary, and one exit 2 -- no fail-fast.
    # First is a transfer sharing a timed trade's day; second is a date-only trade
    # sharing a day with another trade -- both reported through the one mechanism.
    logger = logging.getLogger("ordering_report_fatal")
    collisions = [
        OrderingCollision("AAPL", "USD", dt.date(2024, 6, 10), 1, 0, 1),
        OrderingCollision("VOD", "GBP", dt.date(2024, 7, 1), 2, 1, 0),
    ]
    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        report_ordering_collisions(collisions, logger)

    assert exc.value.code == 2
    messages = [r.getMessage() for r in caplog.records]
    assert any("Unorderable same-day events" in m and "AAPL" in m for m in messages)
    assert any("VOD" in m for m in messages)
    assert any("2 symbol-day(s)" in m for m in messages)


# --- statement input conflicts: detector --------------------------------------------


def _meta_model(*, account=None, period=None):
    """Build an IbkrModel carrying just the Account/Statement metadata under test."""
    rows = []
    if period is not None:
        rows += [
            ["Statement", "Header", "Field Name", "Field Value"],
            ["Statement", "Data", "Period", period],
        ]
    if account is not None:
        rows += [
            ["Account Information", "Header", "Field Name", "Field Value"],
            ["Account Information", "Data", "Account", account],
        ]
    model, _ = IbkrStatementCsvParser().parse_rows(rows)
    return model


def _stmt(path, *, account, period_str):
    """A StatementInput with already-valid identity, the detector's input shape."""
    return StatementInput(
        path, StatementMetadata(account, StatementPeriod.parse(period_str))
    )


def test_detect_statement_conflicts_single_file_is_skipped():
    # A lone statement has nothing to overlap.
    assert (
        detect_statement_input_conflicts(
            [_stmt("a.csv", account="U1", period_str=Y2024)]
        )
        == []
    )


def test_detect_statement_conflicts_disjoint_periods_ok():
    statements = [
        _stmt("a.csv", account="U1", period_str=Y2023),
        _stmt("b.csv", account="U1", period_str=Y2024),
    ]
    assert detect_statement_input_conflicts(statements) == []


def test_detect_statement_conflicts_overlapping_periods():
    statements = [
        _stmt("a.csv", account="U1", period_str=Y2024),
        _stmt("b.csv", account="U1", period_str="June 1, 2024 - December 31, 2024"),
    ]
    problems = detect_statement_input_conflicts(statements)
    assert any(
        "overlapping periods" in p and "a.csv" in p and "b.csv" in p for p in problems
    )


def test_detect_statement_conflicts_same_period_twice():
    # The same file passed twice is the degenerate identical-interval overlap.
    s = _stmt("a.csv", account="U1", period_str=Y2024)
    problems = detect_statement_input_conflicts([s, s])
    assert any("overlapping periods" in p for p in problems)


def test_detect_statement_conflicts_multiple_accounts():
    # Different accounts, even with disjoint periods, are out of scope (co-mingling).
    statements = [
        _stmt("a.csv", account="U1", period_str=Y2023),
        _stmt("b.csv", account="U2", period_str=Y2024),
    ]
    problems = detect_statement_input_conflicts(statements)
    assert any("multiple accounts" in p and "U1" in p and "U2" in p for p in problems)


def test_detect_statement_conflicts_accumulates_every_problem():
    # Three statements, three overlapping pairs: all are returned.
    statements = [
        _stmt(f"{name}.csv", account="U1", period_str=Y2024) for name in ("a", "b", "c")
    ]
    problems = detect_statement_input_conflicts(statements)
    overlaps = [p for p in problems if "overlapping periods" in p]
    assert len(overlaps) == 3  # (a,b), (a,c), (b,c)


# --- statement identity: partition ---------------------------------------------------


def test_partition_statements_all_valid_yields_no_problems():
    models = [
        _meta_model(account="U1", period=Y2023),
        _meta_model(account="U1", period=Y2024),
    ]
    selected, problems = partition_statements_by_metadata(["a.csv", "b.csv"], models)
    assert problems == []
    assert [(s.path, s.metadata.account) for s in selected] == [
        ("a.csv", "U1"),
        ("b.csv", "U1"),
    ]


def test_partition_statements_flags_missing_period():
    models = [
        _meta_model(account="U1", period=Y2023),
        _meta_model(account="U1", period=None),
    ]
    selected, problems = partition_statements_by_metadata(["a.csv", "b.csv"], models)
    assert [s.path for s in selected] == ["a.csv"]
    assert any("b.csv" in p and "missing reporting period" in p for p in problems)


def test_partition_statements_flags_missing_account():
    models = [
        _meta_model(account="U1", period=Y2023),
        _meta_model(account=None, period=Y2024),
    ]
    selected, problems = partition_statements_by_metadata(["a.csv", "b.csv"], models)
    assert [s.path for s in selected] == ["a.csv"]
    assert any("b.csv" in p and "missing account number" in p for p in problems)


def test_partition_statements_flags_unparseable_period():
    models = [
        _meta_model(account="U1", period=Y2023),
        _meta_model(account="U1", period="2024-01-01 to 2024-12-31"),
    ]
    selected, problems = partition_statements_by_metadata(["a.csv", "b.csv"], models)
    assert [s.path for s in selected] == ["a.csv"]
    assert any("b.csv" in p and "statement period" in p.lower() for p in problems)


# --- statement input conflicts: reporter ---------------------------------------------


def test_report_statement_input_conflicts_silent_when_empty(caplog):
    logger = logging.getLogger("conflicts_report_none")
    with caplog.at_level(logging.ERROR):
        report_statement_input_conflicts([], logger)
    assert caplog.records == []


def test_report_statement_input_conflicts_lists_and_exits(caplog):
    logger = logging.getLogger("conflicts_report_fatal")
    problems = [
        "overlapping periods: a.csv and b.csv both cover 2024-06-01 to 2024-12-31.",
        "inputs span multiple accounts (U1, U2); this tool reports one account at a "
        "time.",
    ]
    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        report_statement_input_conflicts(problems, logger)

    assert exc.value.code == 2
    messages = [r.getMessage() for r in caplog.records]
    assert any("overlapping periods" in m for m in messages)
    assert any("multiple accounts" in m for m in messages)
    assert any("single-account, non-overlapping set" in m for m in messages)


# --- statement identity: reporter ----------------------------------------------------


def test_report_invalid_statements_silent_when_empty(caplog):
    logger = logging.getLogger("invalid_report_none")
    with caplog.at_level(logging.ERROR):
        report_invalid_statements([], logger)
    assert caplog.records == []


def test_report_invalid_statements_lists_and_exits(caplog):
    # One ERROR per malformed file, then a summary naming what a valid identity needs,
    # then a single exit 2 -- no fail-fast, mirroring the other boundary reporters.
    logger = logging.getLogger("invalid_report_fatal")
    problems = [
        "a.csv: missing account number (Account Information)",
        "b.csv: Unparseable statement period: '2024-01-01 to 2024-12-31'",
    ]
    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        report_invalid_statements(problems, logger)

    assert exc.value.code == 2
    messages = [r.getMessage() for r in caplog.records]
    assert any("a.csv" in m and "missing account number" in m for m in messages)
    assert any("b.csv" in m for m in messages)
    assert any("missing or malformed identity" in m for m in messages)


# --- run() integration --------------------------------------------------------


_TRADES_HEADER = header_row("Trades", TRADES_COLUMNS)
_TRANSFERS_HEADER = header_row("Transfers", TRANSFERS_COLUMNS)


def _transfer_data(symbol, currency, date, *, direction="In", qty="10"):
    return transfer_data(
        symbol=symbol, currency=currency, date=date, direction=direction, quantity=qty
    )


def _trade(symbol, currency, *, date="2024-01-10, 10:00:00", qty="10"):
    return trade_data(symbol=symbol, currency=currency, datetime_str=date, quantity=qty)


def _write_statement(path, *, currency="USD"):
    # A matched BUY(10) then SELL(10): no gap, so the gap tie-out is a no-op. USD needs
    # an FX table to convert; EUR converts by identity and runs clean to the write step.
    # Carries a valid Account/Period identity so the run clears the identity gate and
    # reaches the condition each dependent test actually exercises.
    rows = _statement_meta_rows("U1", Y2024) + [
        _TRADES_HEADER,
        _trade("AAPL", currency),
        trade_data(
            currency=currency,
            datetime_str="2024-06-10, 10:00:00",
            quantity="-10",
            t_price="110",
            proceeds="1100",
            comm_fee="-1",
            code="C",
            basis="-1001",
            realized="99",
        ),
    ]
    with open(path, "w", newline="", encoding="utf-8") as fp:
        csv.writer(fp).writerows(rows)


def _write_single_sell(path, *, symbol, currency, basis, realized, proceeds="1200"):
    # A lone SELL with no prior BUY: the whole quantity is an unmatched gap. Carries a
    # valid Account/Period identity so the run clears the identity gate and reaches the
    # gap tie-out each dependent test exercises.
    sell = trade_data(
        symbol=symbol,
        currency=currency,
        datetime_str="2024-06-10, 10:00:00",
        quantity="-10",
        t_price="120",
        proceeds=proceeds,
        comm_fee="0",
        code="C",
        basis=basis,
        realized=realized,
    )
    with open(path, "w", newline="", encoding="utf-8") as fp:
        csv.writer(fp).writerows(
            _statement_meta_rows("U1", Y2024) + [_TRADES_HEADER, sell]
        )


def _args(stmt, out_path, *, fx_table=None, auto_fix_sell_gaps=None, dry_run=False):
    return RunOptions(
        inputs=[str(stmt)],
        year=2024,
        fx_table=fx_table,
        locale="EN",
        output=str(out_path),
        auto_fix_sell_gaps=auto_fix_sell_gaps,
        dry_run=dry_run,
    )


def test_process_files_exits_2_when_fx_table_incomplete(tmp_path, out_path, caplog):
    # Non-EUR data with no FX table cannot be converted. A complete table is a
    # precondition, so this aborts with exit 2 and writes no workbook with
    # blank/substituted EUR figures, rather than warning + exit 0.
    stmt = tmp_path / "stmt.csv"
    _write_statement(stmt)

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("Missing FX rate" in r.getMessage() for r in caplog.records)


def test_run_exits_2_on_unorderable_same_day_trades(tmp_path, out_path, caplog):
    # Two EUR trades for the same symbol on the same day -- one stamped date-only (no
    # intraday time), one timestamped -- have no FIFO order in the data. As with a
    # same-day transfer collision, the run must abort (exit 2, no workbook) rather than
    # fabricate an order from the raw Date/Time string. EUR converts by identity, so the
    # run reaches the ordering stage without needing an FX table.
    stmt = tmp_path / "stmt.csv"
    rows = _statement_meta_rows("U1", Y2024) + [
        _TRADES_HEADER,
        trade_data(
            symbol="AAA",
            currency="EUR",
            datetime_str="2024-06-15",
            quantity="10",
            proceeds="-1000",
            comm_fee="0",
            code="O",
        ),
        trade_data(
            symbol="AAA",
            currency="EUR",
            datetime_str="2024-06-15, 09:30:00",
            quantity="-10",
            t_price="110",
            proceeds="1100",
            comm_fee="0",
            code="C",
            basis="-1000",
            realized="100",
        ),
    ]
    with open(stmt, "w", newline="", encoding="utf-8") as fp:
        csv.writer(fp).writerows(rows)

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()


def test_run_does_not_swallow_a_reconciliation_crash(
    tmp_path, out_path, monkeypatch, caplog
):
    # Reconciliation is advisory in what it *finds* (a discrepancy only warns), but it
    # must still *run*: it is the only independent cross-check on our own figures. An
    # unexpected exception inside it is a programming error, not an advisory finding, so
    # it propagates and halts the run (no workbook) rather than being swallowed into one
    # log line under a report that ships unverified. A valid EUR buy+sell reaches the
    # reconciliation stage (no FX table needed); the detector is forced to raise there.
    stmt = tmp_path / "stmt.csv"
    rows = _statement_meta_rows("U1", Y2024) + [
        _TRADES_HEADER,
        trade_data(
            symbol="AAA",
            currency="EUR",
            datetime_str="2024-06-14, 09:30:00",
            quantity="10",
            proceeds="-1000",
            comm_fee="0",
            code="O",
        ),
        trade_data(
            symbol="AAA",
            currency="EUR",
            datetime_str="2024-06-15, 09:30:00",
            quantity="-10",
            t_price="110",
            proceeds="1100",
            comm_fee="0",
            code="C",
            basis="-1000",
            realized="100",
        ),
    ]
    with open(stmt, "w", newline="", encoding="utf-8") as fp:
        csv.writer(fp).writerows(rows)

    def _boom(*args, **kwargs):
        raise RuntimeError("reconciliation bug")

    monkeypatch.setattr(
        "capitangains.pipeline.reconcile_realized_against_ibkr", _boom
    )

    with pytest.raises(RuntimeError, match="reconciliation bug"):
        run(_args(stmt, out_path))

    assert not out_path.exists()


def test_process_files_exits_1_when_fx_table_unreadable(tmp_path, out_path, caplog):
    # A missing or unparseable --fx-table file is a setup failure, not a statement-data
    # defect, so it halts with exit 1 -- a class apart from the curated gates' exit 2 --
    # surfaced as one clean ERROR (explicit, not an emergent raw crash) and no workbook.
    stmt = tmp_path / "stmt.csv"
    _write_statement(stmt)  # USD: needs the FX table to convert
    args = _args(stmt, out_path, fx_table=str(tmp_path / "does_not_exist.csv"))

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(args)

    assert exc.value.code == 1
    assert not out_path.exists()
    assert any(
        "Failed to prepare FX conversion" in r.getMessage() for r in caplog.records
    )


def test_process_files_exits_2_on_malformed_acknowledgment_spec(
    tmp_path, out_path, caplog
):
    # A malformed spec must abort before any file is read (fail fast), with exit 2, no
    # workbook, and an ERROR naming the bad token.
    stmt = tmp_path / "stmt.csv"
    _write_statement(stmt)  # valid statement; only the spec is malformed
    args = _args(stmt, out_path, auto_fix_sell_gaps="BABA@oops")

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(args)

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("BABA@oops" in r.getMessage() for r in caplog.records)


def test_process_files_exits_2_on_malformed_trade_row(
    write_statement, out_path, caplog
):
    # A row-level parse defect (blank Date/Time) must abort cleanly with exit 2, not a
    # raw traceback (exit 1), and write no workbook.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [_TRADES_HEADER, _trade("AAPL", "USD", date="")]
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("Date/Time" in r.getMessage() for r in caplog.records)


def test_process_files_exits_2_on_symbol_currency_violation(
    write_statement, out_path, caplog
):
    # A legitimate-but-rejected data condition (one symbol in two currencies) must abort
    # with exit 2, not a raw traceback.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [_TRADES_HEADER, _trade("ABC", "USD"), _trade("ABC", "EUR")]
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("symbol-currency uniqueness" in r.getMessage() for r in caplog.records)


def test_process_files_exits_2_on_same_day_transfer_trade_collision(
    write_statement, out_path, caplog
):
    # IBKR gives transfers no intraday time, so a transfer landing on the same day as a
    # trade in the SAME symbol cannot be ordered for FIFO. The tool refuses to guess and
    # aborts with exit 2 -- naming the symbol/day -- rather than fabricate an order.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [
            _TRADES_HEADER,
            _trade("AAPL", "USD", date="2024-06-10, 10:00:00"),
            _TRANSFERS_HEADER,
            _transfer_data("AAPL", "USD", "2024-06-10"),
        ]
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any(
        "Unorderable same-day events" in r.getMessage() and "AAPL" in r.getMessage()
        for r in caplog.records
    )


def test_process_files_allows_same_day_transfer_in_a_different_symbol(
    write_statement, out_path, caplog
):
    # The guard is scoped to (symbol, currency): a transfer sharing a day with trades in
    # an UNRELATED symbol is independent for FIFO and must not halt. EUR converts by
    # identity, so the run completes and writes the workbook.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [
            _TRADES_HEADER,
            _trade("AAPL", "EUR", date="2024-01-10, 10:00:00"),
            [
                "Trades",
                "Data",
                "Order",
                "Stocks",
                "EUR",
                "AAPL",
                "2024-06-10, 10:00:00",
                "-10",
                "110",
                "1100",
                "-1",
                "C",
                "-1001",
                "99",
            ],
            _TRANSFERS_HEADER,
            _transfer_data("MSFT", "EUR", "2024-06-10"),
        ]
    )

    with caplog.at_level(logging.ERROR):
        run(_args(stmt, out_path))

    assert out_path.exists()
    assert not any(
        "Unorderable same-day events" in r.getMessage() for r in caplog.records
    )


def test_process_files_transfers_sheet_shows_only_reporting_year(
    write_statement, out_path
):
    # The report is a single-year document, so the Stock Transfers sheet -- like every
    # other category -- shows only args.year. A prior-year transfer is supplied solely
    # to seed FIFO (here SEEDLOT, 2023); it must not surface on the 2024 sheet as if it
    # were a current-year event, while a genuine 2024 transfer (CURRENTXFER) must. Both
    # are IN transfers seeding open lots; all EUR, so the run converts by identity.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [
            _TRANSFERS_HEADER,
            _transfer_data("SEEDLOT", "EUR", "2023-11-01"),
            _transfer_data("CURRENTXFER", "EUR", "2024-03-01"),
        ]
    )

    run(_args(stmt, out_path))

    assert out_path.exists()
    ws = load_workbook(out_path)["Stock Transfers"]
    symbols = {row[1] for row in ws.iter_rows(min_row=2, values_only=True)}
    assert symbols == {"CURRENTXFER"}


def test_process_files_exits_2_on_malformed_dividend_amount(
    write_statement, out_path, caplog
):
    # A present-but-malformed cash-flow value (bad dividend Amount) must abort with
    # exit 2 like trades/SYEP/transfers, not escape as a raw traceback. The
    # skip-incomplete gate only drops rows missing core fields, not malformed amounts.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [
            ["Dividends", "Header", "Currency", "Date", "Description", "Amount"],
            ["Dividends", "Data", "USD", "2024-01-15", "AAPL Dividend", "invalid"],
        ]
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("Amount" in r.getMessage() for r in caplog.records)


def test_process_files_lists_every_extraction_defect(write_statement, out_path, caplog):
    # Accumulate, do not fail fast: a malformed trade row AND a malformed dividend row
    # in the same statement are BOTH reported in one run, then a single exit 2 with no
    # workbook -- so the operator fixes every defect in one pass, not one-per-rerun.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [
            _TRADES_HEADER,
            _trade("AAPL", "USD", date=""),  # bad trade: blank Date/Time
            ["Dividends", "Header", "Currency", "Date", "Description", "Amount"],
            ["Dividends", "Data", "USD", "2024-01-15", "AAPL Dividend", "invalid"],
        ]
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    messages = [r.getMessage() for r in caplog.records]
    assert any("Trades" in m and "Date/Time" in m for m in messages)
    assert any("Dividends" in m and "Amount" in m for m in messages)
    assert any("rejected 2 row(s)" in m for m in messages)


def test_process_files_exits_2_on_unacknowledged_gap(tmp_path, out_path, caplog):
    # A real SELL gap left unacknowledged (no spec) is fatal: a known taxable disposal
    # must never be silently valued at zero cost.
    stmt = tmp_path / "stmt.csv"
    _write_single_sell(
        stmt, symbol="ORPH", currency="USD", basis="-1000", realized="200"
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("Unacknowledged SELL gap" in r.getMessage() for r in caplog.records)


def test_process_files_exits_2_on_orphan_acknowledgment(tmp_path, out_path, caplog):
    # A clean statement (no gaps) with an acknowledgment that matches nothing is fatal:
    # the operator must not carry a stale or mistyped acknowledgment.
    stmt = tmp_path / "stmt.csv"
    _write_statement(stmt)  # fully matched: no gaps
    args = _args(stmt, out_path, auto_fix_sell_gaps="GHOST@2024-01-01")

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(args)

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any(
        "Orphan acknowledgment" in r.getMessage() and "GHOST" in r.getMessage()
        for r in caplog.records
    )


def test_process_files_exits_2_on_acknowledged_gap_with_corrupt_basis(
    tmp_path, out_path, caplog
):
    # An acknowledged gap whose Basis contradicts IBKR's own Realized P/L
    # (Proceeds + Comm + Basis != Realized) is DEFECTIVE; synthesizing from it would
    # fabricate the gain, so the run aborts with exit 2 and writes no workbook. The rich
    # message survives the move to the boundary (mentions both "Basis" and "Realized").
    stmt = tmp_path / "stmt.csv"
    _write_single_sell(
        stmt, symbol="CORRUPT", currency="USD", basis="-99999", realized="200"
    )
    args = _args(stmt, out_path, auto_fix_sell_gaps="CORRUPT@2024-06-10")

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(args)

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any(
        "Basis" in r.getMessage() and "Realized" in r.getMessage()
        for r in caplog.records
    )


def test_process_files_exits_2_on_acknowledged_gap_with_missing_basis(
    tmp_path, out_path, caplog
):
    # An acknowledged gap with no IBKR Basis at all is DEFECTIVE: there is no figure to
    # synthesize from, so the run aborts (exit 2, no workbook) rather than electing a
    # zero basis.
    stmt = tmp_path / "stmt.csv"
    _write_single_sell(stmt, symbol="NOBASIS", currency="USD", basis="", realized="")
    args = _args(stmt, out_path, auto_fix_sell_gaps="NOBASIS@2024-06-10")

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(args)

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("defective IBKR Basis" in r.getMessage() for r in caplog.records)


def test_process_files_synthesizes_acknowledged_gap_and_writes_workbook(
    tmp_path, out_path, caplog
):
    # An acknowledged gap with a valid IBKR Basis is synthesized: the run completes,
    # writes the workbook, and emits the per-lot audit warning so the synthetic basis is
    # never silent. EUR trades need no FX table (identity conversion).
    stmt = tmp_path / "stmt.csv"
    _write_single_sell(
        stmt, symbol="EURGAP", currency="EUR", basis="-1000", realized="200"
    )
    args = _args(stmt, out_path, auto_fix_sell_gaps="EURGAP@2024-06-10")

    with caplog.at_level(logging.WARNING):
        run(args)

    assert out_path.exists()
    assert any("Synthesized residual lot" in r.getMessage() for r in caplog.records)


# --- --dry-run / -n -------------------------------------------------------------------


def test_argparser_accepts_dry_run_short_and_long_flags():
    parser = build_argparser()
    for flag in ("-n", "--dry-run"):
        args = parser.parse_args(["--year", "2024", "stmt.csv", flag])
        assert args.dry_run is True
    # Absent by default: a normal run still writes.
    assert parser.parse_args(["--year", "2024", "stmt.csv"]).dry_run is False


def test_process_files_dry_run_writes_no_workbook_on_clean_input(
    tmp_path, out_path, caplog
):
    # A run that would otherwise succeed writes nothing under --dry-run: it returns
    # normally (exit 0), leaves no file at out_path, and announces the preflight.
    stmt = tmp_path / "stmt.csv"
    _write_statement(stmt, currency="EUR")  # matched, EUR: clean through to the write
    args = _args(stmt, out_path, dry_run=True)

    with caplog.at_level(logging.INFO):
        run(args)  # no SystemExit

    assert not out_path.exists()
    assert any("Dry run" in r.getMessage() for r in caplog.records)


def test_process_files_dry_run_does_not_overwrite_existing_output(tmp_path, out_path):
    # The existing report at out_path is left byte-for-byte untouched: --dry-run never
    # clobbers a prior good workbook.
    stmt = tmp_path / "stmt.csv"
    _write_statement(stmt, currency="EUR")
    out_path.write_bytes(b"prior report")
    args = _args(stmt, out_path, dry_run=True)

    run(args)

    assert out_path.read_bytes() == b"prior report"


def test_process_files_dry_run_still_exits_2_on_defect(
    write_statement, out_path, caplog
):
    # --dry-run validates; it does not suppress failures. A row-level defect still
    # aborts with exit 2 and writes no workbook, exactly as a real run would.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [_TRADES_HEADER, _trade("AAPL", "EUR", date="")]
    )
    args = _args(stmt, out_path, dry_run=True)

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(args)

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("Date/Time" in r.getMessage() for r in caplog.records)


# --- run() multi-file overlap gate --------------------------------------------


def _statement_meta_rows(account, period):
    return statement_meta_rows(account=account, period=period)


def _write_full_statement(
    path, *, account, period, year, currency="EUR", legs="both", realized="99"
):
    # A BUY(10) and/or SELL(10) in `year` plus the Account/Statement metadata the
    # multi-file overlap gate reads. EUR converts by identity, so no FX table is needed.
    # `legs` selects which trade legs to emit: "both" (matched buy then sell), "open"
    # (a lone buy, e.g. a prior-year seeding lot), or "close" (a lone sell). `realized`
    # sets the sell's IBKR Realized P/L, so a test can plant an obviously wrong value
    # to drive a reconciliation mismatch. The defaults reproduce the original pair.
    buy = _trade("AAPL", currency, date=f"{year}-02-10, 10:00:00")
    sell = trade_data(
        currency=currency,
        datetime_str=f"{year}-06-10, 10:00:00",
        quantity="-10",
        t_price="110",
        proceeds="1100",
        comm_fee="-1",
        code="C",
        basis="-1001",
        realized=realized,
    )
    legs_rows = {"both": [buy, sell], "open": [buy], "close": [sell]}[legs]
    rows = _statement_meta_rows(account, period) + [_TRADES_HEADER, *legs_rows]
    with open(path, "w", newline="", encoding="utf-8") as fp:
        csv.writer(fp).writerows(rows)


def _args_multi(stmts, out_path, *, year=2024):
    return RunOptions(
        inputs=[str(s) for s in stmts],
        year=year,
        fx_table=None,
        locale="EN",
        output=str(out_path),
        auto_fix_sell_gaps=None,
        dry_run=False,
    )


def test_process_files_disjoint_statements_write_workbook(tmp_path, out_path):
    # Two single-account statements for adjacent, non-overlapping years merge cleanly
    # and produce a workbook: the gate must not flag a legitimate multi-file run.
    a = tmp_path / "2023.csv"
    b = tmp_path / "2024.csv"
    _write_full_statement(a, account="U1", period=Y2023, year=2023)
    _write_full_statement(b, account="U1", period=Y2024, year=2024)

    run(_args_multi([a, b], out_path))

    assert out_path.exists()


def test_process_files_exits_2_on_overlapping_statements(tmp_path, out_path, caplog):
    # The same period supplied twice would double-count trades into FIFO; the run aborts
    # before merging and writes no workbook.
    a = tmp_path / "2024_a.csv"
    b = tmp_path / "2024_b.csv"
    _write_full_statement(a, account="U1", period=Y2024, year=2024)
    _write_full_statement(b, account="U1", period=Y2024, year=2024)

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args_multi([a, b], out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any("overlapping periods" in r.getMessage() for r in caplog.records)


def test_process_files_multifile_runs_reconciliation(tmp_path, out_path, caplog):
    # The reconciliation now runs for multi-file input: the old single-file guard is
    # gone. Two disjoint self-contained years (mirroring the disjoint-write test)
    # reconcile, so the per-symbol DEBUG trace appears and the "Skipping..." line never
    # does. The check is purely diagnostic, so the workbook is still written.
    a = tmp_path / "2023.csv"
    b = tmp_path / "2024.csv"
    _write_full_statement(a, account="U1", period=Y2023, year=2023)
    _write_full_statement(b, account="U1", period=Y2024, year=2024)

    with caplog.at_level(logging.DEBUG):
        run(_args_multi([a, b], out_path))

    assert out_path.exists()
    messages = [r.getMessage() for r in caplog.records]
    assert not any("Skipping IBKR realized-P/L reconciliation" in m for m in messages)
    assert any(
        r.levelno == logging.DEBUG and r.getMessage().startswith("Reconciliation:")
        for r in caplog.records
    )


def test_process_files_multifile_reconciles_cross_year_lot(tmp_path, out_path, caplog):
    # The case the old guard suppressed: a 2024 sell whose opening lot lives in a
    # 2023 file. Multi-file mode seeds FIFO with the real 2023 buy, so the 2024 sell
    # matches a genuine lot -- no gap, no --auto-fix-sell-gaps, no synthesized basis --
    # and its basis is independent of IBKR. With IBKR's Realized P/L planted at an
    # obviously wrong value, the cross-check a single-file run could never make now
    # catches the disagreement.
    prior = tmp_path / "2023.csv"
    current = tmp_path / "2024.csv"
    _write_full_statement(prior, account="U1", period=Y2023, year=2023, legs="open")
    _write_full_statement(
        current, account="U1", period=Y2024, year=2024, legs="close", realized="9999"
    )

    with caplog.at_level(logging.DEBUG):
        run(_args_multi([prior, current], out_path))

    assert out_path.exists()
    messages = [r.getMessage() for r in caplog.records]
    assert not any("synthesized basis" in m for m in messages)
    assert any("disagree with IBKR realized P/L beyond rounding" in m for m in messages)


# --- run() statement-identity gate --------------------------------------------


def test_process_files_single_file_missing_account_exits_2(tmp_path, out_path, caplog):
    # Identity is validated unconditionally, not only on the multi-file path: a lone
    # statement whose Account is blank halts at the identity gate with exit 2 and no
    # workbook, before any contents are trusted.
    stmt = tmp_path / "stmt.csv"
    _write_full_statement(stmt, account="", period=Y2024, year=2024)

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    messages = [r.getMessage() for r in caplog.records]
    assert any("missing account number" in m for m in messages)
    assert any("missing or malformed identity" in m for m in messages)


def test_process_files_single_file_malformed_period_exits_2(tmp_path, out_path, caplog):
    # A single statement whose Period does not parse is equally fatal at the identity
    # gate -- the precondition is establish identity before trusting contents.
    stmt = tmp_path / "stmt.csv"
    _write_full_statement(
        stmt, account="U1", period="2024-01-01 to 2024-12-31", year=2024
    )

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args(stmt, out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any(
        "missing or malformed identity" in r.getMessage() for r in caplog.records
    )


def test_process_files_single_file_valid_identity_writes_workbook(tmp_path, out_path):
    # The control: a lone statement with a sound Account/Period identity clears the gate
    # and runs through to the workbook. EUR converts by identity, so no FX table needed.
    stmt = tmp_path / "stmt.csv"
    _write_full_statement(stmt, account="U1", period=Y2024, year=2024)

    run(_args(stmt, out_path))

    assert out_path.exists()


def test_process_files_multifile_one_invalid_identity_exits_2(
    tmp_path, out_path, caplog
):
    # In a multi-file set, a single member with an unparseable identity halts the whole
    # run at the identity gate (exit 2, no workbook) -- before the cross-file conflict
    # check, which assumes every period is parseable.
    a = tmp_path / "2023.csv"
    b = tmp_path / "2024.csv"
    _write_full_statement(a, account="U1", period=Y2023, year=2023)
    _write_full_statement(b, account="U1", period="2024-01-01 to 2024-12-31", year=2024)

    with caplog.at_level(logging.ERROR), pytest.raises(SystemExit) as exc:
        run(_args_multi([a, b], out_path))

    assert exc.value.code == 2
    assert not out_path.exists()
    assert any(
        "missing or malformed identity" in r.getMessage() for r in caplog.records
    )


# --- format_reconciliation_sample ----------------------------------------------------


def _recon(symbol, computed, ibkr):
    return SymbolReconciliation(
        symbol=symbol, currency="USD", computed=computed, ibkr=ibkr, n_sells=1
    )


def test_format_reconciliation_sample_lists_all_when_within_cap():
    items = [_recon("AAA", Decimal("1"), Decimal("2"))]
    rendered = format_reconciliation_sample(items)
    assert "showing" not in rendered
    assert "AAA (USD) mine=1 IBKR=2 diff=1" in rendered


def test_format_reconciliation_sample_labels_truncation():
    # More than the cap: the line announces "showing K of N" so a truncated sample is
    # never mistaken for the full set, and emits exactly the cap's worth of entries.
    items = [_recon(f"S{i:02d}", Decimal("1"), Decimal("3")) for i in range(12)]
    rendered = format_reconciliation_sample(items)
    assert rendered.startswith("showing 10 of 12:")
    assert rendered.count("mine=") == 10
    assert "S00" in rendered and "S09" in rendered  # first 10 shown
    assert "S10" not in rendered and "S11" not in rendered  # remainder truncated


def test_report_reconciliation_separates_classes_by_severity(caplog):
    # Sign flips and magnitude gaps each get their own WARNING; synthesized-basis keys
    # are reported apart at INFO. One render, three distinct, correctly-leveled lines.
    report = ReconciliationReport(
        reconciled=[
            _recon("FLIP", Decimal("1000"), Decimal("-2000")),  # sign flip
            _recon("GAP", Decimal("1000"), Decimal("5000")),  # magnitude gap
        ],
        synthetic=[_recon("SYN", Decimal("42"), Decimal("42"))],
        incomplete=[],
    )
    logger = logging.getLogger("recon_render")

    with caplog.at_level(logging.DEBUG, logger="recon_render"):
        report_reconciliation(report, logger)

    leveled = [(r.levelno, r.getMessage()) for r in caplog.records]
    assert any(
        lvl == logging.WARNING and "gain/loss direction" in m and "FLIP" in m
        for lvl, m in leveled
    )
    assert any(
        lvl == logging.WARNING and "beyond rounding" in m and "GAP" in m
        for lvl, m in leveled
    )
    assert any(
        lvl == logging.INFO and "synthesized basis" in m and "SYN" in m
        for lvl, m in leveled
    )


# --- unrecognized-section coverage sweep: detector -----------------------------------


def test_detect_unrecognized_sections_flags_unknown_alongside_consumed():
    # A "Mystery" section (1 subtable, 2 rows) sits next to a consumed "Trades":
    # only the unknown one surfaces, carrying its subtable and row counts.
    model = parse_model(
        section_table("Mystery", ["Col1", "Col2"], ["a", "b"], ["c", "d"])
        + [header_row(SEC_TRADES, TRADES_COLUMNS)]
    )
    assert detect_unrecognized_sections(model) == [UnrecognizedSection("Mystery", 1, 2)]


def test_detect_unrecognized_sections_silent_for_consumed_and_ignored():
    # Trades is consumed; "Open Positions" is allow-listed. Neither is drift, so the
    # sweep returns nothing.
    model = parse_model(
        [
            header_row(SEC_TRADES, TRADES_COLUMNS),
            header_row("Open Positions", ["A", "B"]),
        ]
    )
    assert detect_unrecognized_sections(model) == []


def test_detect_unrecognized_sections_surfaces_a_rename():
    # A renamed consumed section ("Withholding Tax" -> "Withholding Tax Foobar")
    # presents under a key no extractor consumes and no allow-list entry covers, so the
    # new name is exactly what the sweep returns -- the drift a rename would hide.
    model = parse_model(
        section_table(
            "Withholding Tax Foobar",
            WITHHOLDING_COLUMNS,
            ["USD", "2024-01-15", "AAPL Cash Dividend - US Tax", "-1.50", ""],
        )
    )
    assert detect_unrecognized_sections(model) == [
        UnrecognizedSection("Withholding Tax Foobar", 1, 1)
    ]


def test_detect_unrecognized_sections_includes_header_only_section():
    # An unknown section with a header but no data rows is still drift; it is returned
    # with row_count == 0 rather than dropped for being empty.
    model = parse_model([header_row("Mystery", ["Col1", "Col2"])])
    assert detect_unrecognized_sections(model) == [UnrecognizedSection("Mystery", 1, 0)]


def test_detect_unrecognized_sections_sums_rows_across_subtables():
    # Two subtables of one unknown section (as a multi-file merge would union) collapse
    # to a single finding whose subtable_count is 2 and row_count sums both (1 + 2).
    model = parse_model(
        section_table("Mystery", ["C"], ["a"])
        + section_table("Mystery", ["C"], ["b"], ["c"])
    )
    assert detect_unrecognized_sections(model) == [UnrecognizedSection("Mystery", 2, 3)]


def test_detect_unrecognized_sections_returns_names_sorted():
    # Two unknown sections come back ordered by name (order=True, name leads), so a
    # report over them is deterministic regardless of statement order.
    model = parse_model([header_row("Zebra", ["C"]), header_row("Alpha", ["C"])])
    assert [s.name for s in detect_unrecognized_sections(model)] == ["Alpha", "Zebra"]


# --- unrecognized-section coverage sweep: reporter -----------------------------------


def test_report_unrecognized_sections_silent_when_empty(caplog):
    logger = logging.getLogger("unrecognized_report_none")
    with caplog.at_level(logging.WARNING):
        report_unrecognized_sections([], logger)
    assert caplog.records == []


def test_report_unrecognized_sections_warns_without_exiting(caplog):
    # The contract that separates this from every other reporter here: it WARNs and
    # returns, never raises SystemExit. One WARNING names the section and its counts
    # with the "renamed or added" guidance, followed by the summary WARNING.
    logger = logging.getLogger("unrecognized_report_warn")
    with caplog.at_level(logging.WARNING):
        report_unrecognized_sections([UnrecognizedSection("Mystery", 1, 2)], logger)

    assert all(r.levelno == logging.WARNING for r in caplog.records)
    messages = [r.getMessage() for r in caplog.records]
    assert any(
        "Mystery" in m
        and "1 subtable(s)" in m
        and "2 row(s)" in m
        and "renamed or added" in m
        for m in messages
    )
    assert any("1 statement section(s) present" in m for m in messages)


# --- Quadro 8A unattributed income: detector + reporter ------------------------------


def _q8a(kind, country, gross="10.00", tax="1.50"):
    return Quadro8ALine(kind, country, Decimal(gross), Decimal(tax))


def test_detect_unattributed_income_flags_empty_country_lines():
    lines = [
        _q8a(IncomeKind.DIVIDEND, "US"),
        _q8a(IncomeKind.DIVIDEND, ""),
        _q8a(IncomeKind.PIL, ""),
        _q8a(IncomeKind.INTEREST, "IE"),
    ]
    assert detect_unattributed_income(lines) == [
        _q8a(IncomeKind.DIVIDEND, ""),
        _q8a(IncomeKind.PIL, ""),
    ]


def test_detect_unattributed_income_silent_when_every_line_has_a_country():
    lines = [_q8a(IncomeKind.DIVIDEND, "US"), _q8a(IncomeKind.INTEREST, "IE")]
    assert detect_unattributed_income(lines) == []


def test_report_unattributed_income_silent_when_empty(caplog):
    logger = logging.getLogger("unattributed_report_none")
    with caplog.at_level(logging.WARNING):
        report_unattributed_income([], logger)
    assert caplog.records == []


def test_report_unattributed_income_warns_without_exiting(caplog):
    # The contract that separates this from the fail-closed reporters: it WARNs and
    # returns, never SystemExit. The gross/tax figures are correct; only the
    # source-country label is missing. One WARNING names the line, then the summary.
    logger = logging.getLogger("unattributed_report_warn")
    line = Quadro8ALine(IncomeKind.DIVIDEND, "", Decimal("9.00"), Decimal("1.35"))
    with caplog.at_level(logging.WARNING):
        report_unattributed_income([line], logger)

    assert all(r.levelno == logging.WARNING for r in caplog.records)
    messages = [r.getMessage() for r in caplog.records]
    assert any(
        "DIVIDEND" in m and "E11" in m and "9.00" in m and "1.35" in m for m in messages
    )
    assert any("1 Quadro 8A line(s)" in m for m in messages)


# --- Quadro 8A orphaned foreign tax: detector + reporter -----------------------------


def test_detect_orphaned_foreign_tax_flags_tax_with_zero_gross():
    lines = [
        _q8a(IncomeKind.DIVIDEND, "US", gross="10.00", tax="1.50"),  # matched
        _q8a(IncomeKind.DIVIDEND, "US", gross="0.00", tax="1.80"),  # orphaned tax
        _q8a(IncomeKind.PIL, "US", gross="3.00", tax="0.00"),  # gross-only, fine
    ]
    assert detect_orphaned_foreign_tax(lines) == [
        _q8a(IncomeKind.DIVIDEND, "US", gross="0.00", tax="1.80"),
    ]


def test_detect_orphaned_foreign_tax_silent_when_every_taxed_line_has_gross():
    lines = [
        _q8a(IncomeKind.DIVIDEND, "US", gross="10.00", tax="1.50"),
        _q8a(IncomeKind.INTEREST, "IE", gross="2.00", tax="0.00"),
    ]
    assert detect_orphaned_foreign_tax(lines) == []


def test_report_orphaned_foreign_tax_silent_when_empty(caplog):
    logger = logging.getLogger("orphaned_tax_report_none")
    with caplog.at_level(logging.WARNING):
        report_orphaned_foreign_tax([], logger)
    assert caplog.records == []


def test_report_orphaned_foreign_tax_warns_without_exiting(caplog):
    # Sibling contract to report_unattributed_income: WARNs and returns, never exits.
    # The tax figure is correct; it just has no gross income line to sit against.
    logger = logging.getLogger("orphaned_tax_report_warn")
    line = Quadro8ALine(IncomeKind.DIVIDEND, "US", Decimal("0.00"), Decimal("1.80"))
    with caplog.at_level(logging.WARNING):
        report_orphaned_foreign_tax([line], logger)

    assert all(r.levelno == logging.WARNING for r in caplog.records)
    messages = [r.getMessage() for r in caplog.records]
    assert any(
        "DIVIDEND" in m and "E11" in m and "US" in m and "1.80" in m for m in messages
    )
    assert any("1 Quadro 8A line(s)" in m for m in messages)


# --- unrecognized-section coverage sweep: sync guards --------------------------------


def test_consumed_sections_matches_the_six_extractor_constants():
    # CONSUMED_SECTIONS is built from the per-extractor constants; pin the membership
    # so a seventh extractor (or a dropped one) cannot silently desync the set.
    assert {
        SEC_TRADES,
        SEC_DIVIDENDS,
        SEC_INTEREST,
        SEC_WITHHOLDING,
        SEC_TRANSFERS,
        SEC_SYEP,
    } == CONSUMED_SECTIONS


def test_syep_section_constant_matches_fixture():
    # The SYEP section name is long and exact; tie the production constant to the
    # fixture column-set's section so a typo in either is caught.
    assert SEC_SYEP == SYEP_SECTION


def test_consumed_and_ignored_sections_are_disjoint():
    # A name cannot be both consumed and allow-listed; an overlap would make the sweep's
    # partition incoherent.
    assert CONSUMED_SECTIONS.isdisjoint(IGNORED_SECTIONS)


def test_run_warns_on_unrecognized_section_and_still_writes(
    write_statement, out_path, caplog
):
    # End to end: a clean, matched EUR statement that also carries an unrecognized
    # "Mystery" section. The sweep is warn-only, so the workbook is still written and a
    # WARNING names the section -- it never gates the run.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [
            _TRADES_HEADER,
            _trade("AAPL", "EUR"),
            trade_data(
                currency="EUR",
                datetime_str="2024-06-10, 10:00:00",
                quantity="-10",
                t_price="110",
                proceeds="1100",
                comm_fee="-1",
                code="C",
                basis="-1001",
                realized="99",
            ),
        ]
        + section_table("Mystery", ["Col1", "Col2"], ["a", "b"])
    )

    with caplog.at_level(logging.WARNING):
        run(_args(stmt, out_path))

    assert out_path.exists()
    assert any(
        r.levelno == logging.WARNING and "Mystery" in r.getMessage()
        for r in caplog.records
    )


def test_run_warns_on_uncountried_dividend_and_still_writes(
    write_statement, out_path, caplog
):
    # End to end: a clean EUR statement (matched trade) plus an EUR dividend whose
    # description carries no ISIN. The income folds into a Quadro 8A line with no source
    # country; the check is warn-only, so the workbook is still written and a WARNING
    # names it -- it never gates the run.
    stmt = write_statement(
        _statement_meta_rows("U1", Y2024)
        + [
            _TRADES_HEADER,
            _trade("AAPL", "EUR"),
            trade_data(
                currency="EUR",
                datetime_str="2024-06-10, 10:00:00",
                quantity="-10",
                t_price="110",
                proceeds="1100",
                comm_fee="-1",
                code="C",
                basis="-1001",
                realized="99",
            ),
            ["Dividends", "Header", "Currency", "Date", "Description", "Amount"],
            ["Dividends", "Data", "EUR", "2024-03-15", "Mystery Dividend", "50"],
        ]
    )

    with caplog.at_level(logging.WARNING):
        run(_args(stmt, out_path))

    assert out_path.exists()
    assert any(
        r.levelno == logging.WARNING
        and "no identifiable source country" in r.getMessage()
        for r in caplog.records
    )
