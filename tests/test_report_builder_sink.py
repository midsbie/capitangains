import datetime as dt
from decimal import Decimal
from typing import Any

import pytest
from openpyxl import load_workbook

from capitangains.conv import Currency
from capitangains.reporting import detect_symbol_currency_violations
from capitangains.reporting.extract import (
    DividendRow,
    InterestRow,
    SyepInterestRow,
    WithholdingRow,
)
from capitangains.reporting.fifo_domain import RealizedLine, SellMatchLeg
from capitangains.reporting.i18n import LABELS, labels_for
from capitangains.reporting.quadro_8a import IncomeKind, Quadro8ALine
from capitangains.reporting.report_builder import ReportBuilder
from capitangains.reporting.report_sink import (
    _REALIZED_SPEC,
    ExcelReportSink,
    _anexo_j_rows,
    _gap_status,
)
from tests.support import convert, make_fx, realized_line, trade_row


def _realized(
    symbol: str,
    currency: str,
    sell_date: dt.date,
    legs: list[dict[str, Any]],
    *,
    has_gap: bool = False,
    gap_fixed: bool = False,
):
    return realized_line(
        symbol=symbol,
        currency=currency,
        sell_date=sell_date,
        legs=legs,
        has_gap=has_gap,
        gap_fixed=gap_fixed,
    )


def test_report_builder_add_realized_accumulates_symbol_totals():
    rb = ReportBuilder(year=2024)
    legs = [
        {
            "buy_date": dt.date(2023, 1, 1),
            "qty": Decimal("5"),
            "alloc_cost_ccy": Decimal("40"),
        }
    ]
    rl1 = _realized("ABC", "USD", dt.date(2024, 1, 1), legs)
    rl2 = _realized("ABC", "USD", dt.date(2024, 2, 1), legs)
    rb.add_realized(rl1)
    rb.add_realized(rl2)

    totals = rb.symbol_totals["ABC"]
    usd = totals.by_currency[Currency("USD")]
    assert usd.realized == rl1.realized_pl_ccy + rl2.realized_pl_ccy
    assert usd.proceeds == rl1.sell_net_ccy + rl2.sell_net_ccy


def test_multi_currency_same_symbol_detected():
    """Same symbol in multiple currencies is a detected violation."""
    trades = [
        trade_row(symbol="ABC", currency=Currency("USD")),
        trade_row(symbol="ABC", currency=Currency("EUR")),
    ]
    assert detect_symbol_currency_violations(trades, []) == {
        "ABC": frozenset({Currency("USD"), Currency("EUR")})
    }


def test_convert_eur_leaves_line_unconverted_when_any_rate_missing():
    """A missing buy- or sell-date rate leaves the entire line unconverted and records
    every unresolved (date, currency); no substitution of another date's rate."""
    rb = ReportBuilder(year=2024)
    # Sell-date rate present, but the first leg's buy date predates the table, missing.
    usd_legs: list[dict[str, Any]] = [
        {
            "buy_date": dt.date(2023, 6, 1),
            "qty": Decimal("5"),
            "alloc_cost_ccy": Decimal("40"),
        },
        {  # zero-cost gap leg: no buy date, cost 0
            "buy_date": None,
            "qty": Decimal("5"),
            "alloc_cost_ccy": Decimal("0"),
        },
    ]
    rl_usd = _realized("USD1", "USD", dt.date(2024, 3, 1), usd_legs)
    # GBP entirely absent from the table, so even the sell-date rate is missing.
    gbp_legs = [
        {
            "buy_date": dt.date(2024, 2, 1),
            "qty": Decimal("2"),
            "alloc_cost_ccy": Decimal("20"),
        }
    ]
    rl_gbp = _realized("GBP1", "GBP", dt.date(2024, 1, 10), gbp_legs)
    rb.add_realized(rl_usd)
    rb.add_realized(rl_gbp)

    rb.convert_eur(make_fx({("USD", "2024-03-01"): Decimal("0.9")}))

    # Neither line is converted: the USD line because one leg's buy-date rate is missing
    # (even though its sell-date rate was available), the GBP line for want of any rate.
    assert rb.converted_lines == []
    # Every unresolved lookup is recorded for the CLI to abort on.
    assert rb.fx_missing == {
        (dt.date(2023, 6, 1), Currency("USD")),
        (dt.date(2024, 1, 10), Currency("GBP")),
        (dt.date(2024, 2, 1), Currency("GBP")),
    }


def test_convert_eur_zero_sell_qty_skips_proceeds_allocation():
    rb = ReportBuilder(year=2024)
    zero_qty_rl = RealizedLine(
        symbol="ZQ",
        currency=Currency("USD"),
        sell_date=dt.date(2024, 4, 1),
        sell_qty=Decimal("0"),
        sell_gross_ccy=Decimal("0"),
        sell_comm_ccy=Decimal("0"),
        sell_net_ccy=Decimal("0"),
        legs=[
            SellMatchLeg(
                buy_date=None,
                qty=Decimal("0"),
                alloc_cost_ccy=Decimal("0"),
            )
        ],
        realized_pl_ccy=Decimal("0"),
    )
    rb.add_realized(zero_qty_rl)
    rb.convert_eur(make_fx({("USD", "2024-04-01"): Decimal("0.9")}))
    assert not rb.fx_missing
    (converted,) = rb.converted_lines
    assert converted.legs[0].proceeds_share_eur == Decimal("0")


def test_report_builder_income_conversion():
    rb = ReportBuilder(year=2024)
    rb.set_dividends(
        [
            DividendRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 1),
                description="Div USD",
                amount=Decimal("10"),
            ),
            DividendRow(
                currency=Currency("EUR"),
                date=dt.date(2024, 1, 2),
                description="Div EUR",
                amount=Decimal("5"),
            ),
        ]
    )
    rb.set_withholding(
        [
            WithholdingRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 1),
                description="Tax",
                amount=Decimal("-2"),
                code="",
                type="",
                country="",
            ),
        ]
    )
    rb.set_syep_interest(
        [
            SyepInterestRow(
                currency=Currency("USD"),
                value_date=dt.date(2024, 1, 1),
                symbol="SYEP",
                start_date=None,
                quantity=Decimal("-1"),
                collateral_amount=Decimal("0"),
                market_rate_pct=Decimal("0"),
                customer_rate_pct=Decimal("0"),
                interest_paid=Decimal("1"),
                code="",
            )
        ]
    )

    fx = make_fx({("USD", "2024-01-01"): Decimal("0.9")})
    rb.convert_eur(fx)

    assert rb.dividends[0].amount_eur == Decimal("9.00")
    assert rb.dividends[1].amount_eur == Decimal("5.00")
    assert rb.withholding[0].amount_eur == Decimal("-1.80")
    assert rb.syep_interest[0].interest_paid_eur == Decimal("0.90")


def _income_report(broker_country: str = "IE") -> ReportBuilder:
    """A converted report carrying a US cash dividend (+US tax), a US payment in lieu,
    and a EUR interest line, used by the Quadro 8A builder and sink tests.
    """
    rb = ReportBuilder(year=2024, broker_country=broker_country)
    rb.set_dividends(
        [
            DividendRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 1),
                description="PACW(US6952631033) Cash Dividend",
                amount=Decimal("10"),
            ),
            DividendRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 2),
                description="PACW(US6952631033) Payment in Lieu of Dividend",
                amount=Decimal("4"),
            ),
        ]
    )
    rb.set_interest(
        [
            InterestRow(
                currency=Currency("EUR"),
                date=dt.date(2024, 1, 3),
                description="EUR Credit Interest for Jan-2024",
                amount=Decimal("2"),
            ),
        ]
    )
    rb.set_withholding(
        [
            WithholdingRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 1),
                description="PACW(US6952631033) Cash Dividend",
                amount=Decimal("-1.50"),
                code="",
                type="Dividend",
                country="US",
            ),
        ]
    )
    rb.convert_eur(
        make_fx(
            {
                ("USD", "2024-01-01"): Decimal("0.9"),
                ("USD", "2024-01-02"): Decimal("0.9"),
            }
        )
    )
    return rb


def test_report_builder_groups_quadro_8a_income():
    rb = _income_report(broker_country="IE")
    assert rb.quadro_8a == [
        Quadro8ALine(IncomeKind.DIVIDEND, "US", Decimal("9.00"), Decimal("1.35")),
        Quadro8ALine(IncomeKind.PIL, "US", Decimal("3.60"), Decimal("0.00")),
        Quadro8ALine(IncomeKind.INTEREST, "IE", Decimal("2.00"), Decimal("0.00")),
    ]

    # The interest source country is the injected jurisdiction, not a constant.
    rb_lu = _income_report(broker_country="LU")
    interest = [line for line in rb_lu.quadro_8a if line.income_code == "E21"]
    assert [line.country for line in interest] == ["LU"]


def test_quadro_8a_sheet_round_trip(tmp_path):
    rb = _income_report()
    out_path = tmp_path / "quadro_8a.xlsx"
    ExcelReportSink(out_path=out_path, locale="EN").write(rb)

    ws = load_workbook(out_path)["Annex J Box 8A Income"]
    grid = [
        [(cell.value, cell.number_format) for cell in row] for row in ws.iter_rows()
    ]

    eur = "€#,##0.00"
    assert grid == [
        [
            ("Income Code", "General"),
            ("Type", "General"),
            ("Source Country", "General"),
            ("Gross Income (EUR)", "General"),
            ("Foreign Tax (EUR)", "General"),
        ],
        [
            ("E11", "General"),
            ("Dividend", "General"),
            ("US", "General"),
            (9.0, eur),
            (1.35, eur),
        ],
        [
            ("E11", "General"),
            ("Payment in Lieu", "General"),
            ("US", "General"),
            (3.6, eur),
            (0, eur),
        ],
        [
            ("E21", "General"),
            ("Interest", "General"),
            ("IE", "General"),
            (2.0, eur),
            (0, eur),
        ],
    ]


def test_excel_report_sink_handles_empty_report(tmp_path):
    rb = ReportBuilder(year=2024)
    out_path = tmp_path / "report.xlsx"
    sink = ExcelReportSink(out_path=out_path, locale="EN")
    sink.write(rb)

    wb = load_workbook(out_path)
    assert set(wb.sheetnames) >= {
        "Trading Totals",
        "Realized Trades",
        "Per Symbol Summary",
    }


def test_excel_report_sink_serializes_legs(tmp_path):
    rb = ReportBuilder(year=2024)
    leg = {
        "buy_date": dt.date(2023, 1, 1),
        "qty": Decimal("5"),
        "alloc_cost_ccy": Decimal("40"),
    }
    rl = _realized("ABC", "USD", dt.date(2024, 1, 1), [leg])
    rb.add_realized(rl)
    rb.convert_eur(
        make_fx(
            {
                ("USD", "2023-01-01"): Decimal("0.9"),
                ("USD", "2024-01-01"): Decimal("0.9"),
            }
        )
    )

    out_path = tmp_path / "report_with_leg.xlsx"
    sink = ExcelReportSink(out_path=out_path, locale="EN")
    sink.write(rb)

    wb = load_workbook(out_path)
    ws = wb["Realized Trades"]
    legs_json = ws.cell(row=2, column=15).value
    assert isinstance(legs_json, str) and '"buy_date": "2023-01-01"' in legs_json


def test_gap_status_reflects_basis_provenance():
    leg = {
        "buy_date": dt.date(2023, 1, 1),
        "qty": Decimal("5"),
        "alloc_cost_ccy": Decimal("40"),
    }
    clean = _realized("AAA", "USD", dt.date(2024, 1, 1), [leg])
    synth = _realized(
        "BBB", "USD", dt.date(2024, 1, 2), [leg], has_gap=True, gap_fixed=True
    )
    zero = _realized("CCC", "USD", dt.date(2024, 1, 3), [leg], has_gap=True)

    assert _gap_status(clean) == ""
    assert _gap_status(synth) == "synthesized from Basis"
    assert _gap_status(zero) == "zero-cost gap"


def test_realized_sheet_flags_synthesized_basis(tmp_path):
    rb = ReportBuilder(year=2024)
    leg = {
        "buy_date": dt.date(2023, 1, 1),
        "qty": Decimal("5"),
        "alloc_cost_ccy": Decimal("40"),
    }
    rb.add_realized(
        _realized(
            "SYN", "USD", dt.date(2024, 1, 1), [leg], has_gap=True, gap_fixed=True
        )
    )
    rb.convert_eur(
        make_fx(
            {
                ("USD", "2023-01-01"): Decimal("0.9"),
                ("USD", "2024-01-01"): Decimal("0.9"),
            }
        )
    )

    out_path = tmp_path / "synth.xlsx"
    ExcelReportSink(out_path=out_path, locale="EN").write(rb)

    ws = load_workbook(out_path)["Realized Trades"]
    col = ws.max_column  # "Basis Status" is appended last
    assert ws.cell(row=1, column=col).value == "Basis Status"
    assert ws.cell(row=2, column=col).value == "synthesized from Basis"


def test_excel_report_sink_sorts_dividends_by_description(tmp_path):
    rb = ReportBuilder(year=2024)
    rb.set_dividends(
        [
            DividendRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 2),
                description="Zulu",
                amount=Decimal("2"),
            ),
            DividendRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 1),
                description="Alpha",
                amount=Decimal("1"),
            ),
        ]
    )

    out_path = tmp_path / "dividends_sorted.xlsx"
    sink = ExcelReportSink(out_path=out_path, locale="EN")
    sink.write(rb)

    wb = load_workbook(out_path)
    ws = wb["Dividends"]
    descriptions = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
    descriptions_str = [str(d) for d in descriptions]
    assert descriptions_str == sorted(descriptions_str)


def test_excel_report_sink_sorts_account_interest(tmp_path):
    rb = ReportBuilder(year=2024)
    rb.set_interest(
        [
            InterestRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 2),
                description="Zulu",
                amount=Decimal("2"),
            ),
            InterestRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 1),
                description="Alpha",
                amount=Decimal("1"),
            ),
        ]
    )

    out_path = tmp_path / "interest_sorted.xlsx"
    sink = ExcelReportSink(out_path=out_path, locale="EN")
    sink.write(rb)

    wb = load_workbook(out_path)
    ws = wb["Account Interest"]
    descriptions = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
    descriptions_str = [str(d) for d in descriptions]
    assert descriptions_str == sorted(descriptions_str)


@pytest.mark.parametrize(
    "currency, fx_rates",
    [
        ("EUR", None),
        (
            "USD",
            {
                ("USD", "2024-06-15"): Decimal("0.9"),  # sell date
                ("USD", "2023-01-01"): Decimal("0.9"),  # buy dates
                ("USD", "2023-01-02"): Decimal("0.9"),
                ("USD", "2023-01-03"): Decimal("0.9"),
            },
        ),
    ],
    ids=["eur_native", "fx_converted"],
)
def test_proceeds_allocation_sums_to_sell_net_eur(currency, fx_rates):
    """Proceeds split across 3 equal legs must sum exactly to sell_net_eur."""
    legs = [
        {
            "buy_date": dt.date(2023, 1, i),
            "qty": Decimal("10"),
            "alloc_cost_ccy": Decimal("30"),
        }
        for i in range(1, 4)
    ]
    rl = _realized("XYZ", currency, dt.date(2024, 6, 15), legs)
    # sell_net_ccy = 100 (from _realized helper); 100 / 3 is non-terminating

    rb = ReportBuilder(year=2024)
    rb.add_realized(rl)
    fx = make_fx(fx_rates) if fx_rates else None
    rb.convert_eur(fx)

    (converted,) = rb.converted_lines
    shares = [cleg.proceeds_share_eur for cleg in converted.legs]
    assert sum(shares, Decimal("0")) == converted.sell_net_eur


def test_excel_report_sink_sorts_withholding(tmp_path):
    rb = ReportBuilder(year=2024)
    rb.set_withholding(
        [
            WithholdingRow(
                currency=Currency("USD"),
                date=dt.date(2024, 1, 3),
                description="Bravo",
                amount=Decimal("-2"),
                code="",
                type="",
                country="",
            ),
            WithholdingRow(
                currency=Currency("EUR"),
                date=dt.date(2024, 1, 1),
                description="Zulu",
                amount=Decimal("-1"),
                code="",
                type="",
                country="",
            ),
            WithholdingRow(
                currency=Currency("EUR"),
                date=dt.date(2024, 1, 2),
                description="Alpha",
                amount=Decimal("-1.5"),
                code="",
                type="",
                country="",
            ),
        ]
    )

    out_path = tmp_path / "withholding_sorted.xlsx"
    sink = ExcelReportSink(out_path=out_path, locale="EN")
    sink.write(rb)

    wb = load_workbook(out_path)
    ws = wb["Withholding Tax"]
    rows = [
        (ws.cell(row=i, column=2).value, ws.cell(row=i, column=3).value)
        for i in range(2, ws.max_row + 1)
    ]
    assert rows == sorted(rows, key=lambda r: (r[0], r[1]))


def test_every_label_field_defines_both_locales():
    """Locale parity: each field must carry exactly the same set of locales.

    The canonical label table co-locates the translations per field precisely so a key
    can never exist in one language alone (separate per-locale dicts could diverge,
    surfacing only as a write-time KeyError in the affected locale). Deriving the locale
    set from the data keeps this honest if a third locale is ever added.
    """
    field_locale_sets = {
        (section, field): frozenset(translations)
        for section, fields in LABELS.items()
        for field, translations in fields.items()
    }
    assert field_locale_sets, "label table is empty"

    expected = {"EN", "PT"}
    diverging = {
        key: sorted(locs)
        for key, locs in field_locale_sets.items()
        if set(locs) != expected
    }
    assert not diverging, f"fields not covering exactly {sorted(expected)}: {diverging}"

    # No empty translations slip through.
    blank = [
        (section, field, loc)
        for section, fields in LABELS.items()
        for field, translations in fields.items()
        for loc, text in translations.items()
        if not text.strip()
    ]
    assert not blank, f"blank label text: {blank}"


def test_pt_projection_renders_portuguese_labels():
    """PT had no text coverage before the unification; pin a representative sample.

    Includes the Anexo-J realized-P/L header, whose PT text carries a non-breaking
    hyphen (U+2011) that must survive byte-for-byte.
    """
    pt = labels_for("PT")
    assert pt["sheet"]["summary"] == "Totais de Operações"
    assert pt["realized"]["ticker"] == "Símbolo"
    assert pt["anexo_j"]["pl_eur"] == "Mais/menos-valia (EUR)"
    # Any unrecognized locale falls back to PT (the report's default).
    assert labels_for("XX") == pt


# Allocated cost must reach the sheets cent-quantized regardless of currency.  FIFO
# basis allocation (round_cost_piece) emits 8-decimal residuals: 100 * 1/3 is stored as
# 33.33333333. convert_eur quantizes each leg's EUR alloc to cents on every path (an FX
# rate, or EUR-native at rate 1); the trade-currency columns instead sum the raw pieces,
# quantized only at the sink's _MoneyColumn boundary.
_BASIS_RESIDUAL = Decimal("33.33333333")  # == round_cost_piece(Decimal("100"), 1, 3)

_RESIDUAL_LEG = {
    "buy_date": dt.date(2024, 1, 10),
    "qty": Decimal("1"),
    "alloc_cost_ccy": _BASIS_RESIDUAL,
}


def _anexo_j_alloc_eur(currency: str, fx) -> Decimal:
    """The single Anexo J row's stored EUR acquisition cost for a one-leg sell."""
    rb = ReportBuilder(year=2024)
    rb.add_realized(
        realized_line(
            symbol="ACME",
            currency=currency,
            sell_date=dt.date(2024, 6, 15),
            sell_gross_ccy="50",
            sell_net_ccy="50",
            legs=[_RESIDUAL_LEG],
        )
    )
    rb.convert_eur(fx)
    (row,) = _anexo_j_rows(rb)
    return row.alloc_eur


def test_anexo_j_eur_alloc_cost_is_cent_quantized_like_the_fx_path():
    # Control: a non-EUR lot priced at exactly 1.0 is quantized to cents by the FX path.
    usd = _anexo_j_alloc_eur(
        "USD",
        make_fx(
            {
                ("USD", "2024-01-10"): Decimal("1"),
                ("USD", "2024-06-15"): Decimal("1"),
            }
        ),
    )
    assert usd == Decimal("33.33")

    # The economically identical EUR-native lot reaches the filer-facing Anexo J sheet
    # at the same cent-exact 33.33, not leaked as the raw 8-decimal residual.
    assert _anexo_j_alloc_eur("EUR", None) == usd


def test_realized_sheet_alloc_tcy_cell_is_cent_quantized():
    # The trade-currency allocated-cost column sums raw 8-decimal leg pieces; the cell
    # openpyxl persists must still be cent-exact, not the sub-cent value the money
    # format would hide. float(Decimal("33.33")) is the exact double of the 33.33
    # literal, so this equality is precise (the raw 33.33333333 cell would not pass).
    rl = realized_line(
        symbol="ACME",
        currency=Currency("EUR"),
        sell_date=dt.date(2024, 6, 15),
        legs=[_RESIDUAL_LEG],
    )
    (alloc_tcy,) = [c for c in _REALIZED_SPEC.columns if c.header_key == "alloc_tcy"]
    assert alloc_tcy.cell_value(convert(rl)) == float(Decimal("33.33"))
