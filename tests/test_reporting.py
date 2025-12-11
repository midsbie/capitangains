import datetime as dt
from decimal import Decimal

from openpyxl import load_workbook

from capitangains.model.ibkr import IbkrStatementCsvParser
from capitangains.reporting.extract import (
    DividendRow,
    SyepInterestRow,
    WithholdingRow,
    parse_interest,
    parse_syep_interest_details,
)
from capitangains.reporting.fifo import RealizedLine
from capitangains.reporting.fx import FxTable
from capitangains.reporting.report_builder import ReportBuilder
from capitangains.reporting.report_sink import ExcelReportSink


def make_fx(rates: dict[tuple[str, str], Decimal]) -> FxTable:
    ft = FxTable()
    # rates: {(currency, yyyy-mm-dd): eur_per_unit}
    for (ccy, d), v in rates.items():
        c = ccy.upper()
        ft.data[c][d] = v
    for c, m in ft.data.items():
        ft.date_index[c] = sorted(m.keys())
    return ft


def test_syep_interest_parsing_excludes_totals():
    rows = [
        [
            "Stock Yield Enhancement Program Securities Lent Interest Details",
            "Header",
            "Currency",
            "Value Date",
            "Symbol",
            "Start Date",
            "Quantity",
            "Collateral Amount",
            "Market-based Rate (%)",
            "Interest Rate on Customer Collateral (%)",
            "Interest Paid to Customer",
            "Code",
        ],
        [
            "Stock Yield Enhancement Program Securities Lent Interest Details",
            "Data",
            "USD",
            "2024-06-10",
            "BANC",
            "2024-06-10",
            "-216",
            "3024",
            "0.1",
            "0.05",
            "0.01",
            "Po",
        ],
        [
            "Stock Yield Enhancement Program Securities Lent Interest Details",
            "Data",
            "Total",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "1.73",
            "",
        ],
    ]
    parser = IbkrStatementCsvParser()
    model, _ = parser.parse_rows(rows)
    parsed = parse_syep_interest_details(model)
    assert len(parsed) == 1
    r = parsed[0]
    assert r.currency == "USD"
    assert r.symbol == "BANC"
    assert r.value_date is not None and r.value_date.isoformat() == "2024-06-10"
    assert r.interest_paid == Decimal("0.01")


def test_interest_parsing_excludes_totals():
    rows = [
        ["Interest", "Header", "Currency", "Date", "Description", "Amount"],
        [
            "Interest",
            "Data",
            "EUR",
            "2024-02-05",
            "EUR Credit Interest for Jan-2024",
            "139.06",
        ],
        ["Interest", "Data", "Total", "", "", "1479.06"],
        ["Interest", "Data", "Total in EUR", "", "", "1479.06"],
    ]
    model, _ = IbkrStatementCsvParser().parse_rows(rows)
    parsed = parse_interest(model)
    assert len(parsed) == 1
    assert parsed[0].currency == "EUR"
    assert parsed[0].amount == Decimal("139.06")


def test_convert_eur_for_income_rows(tmp_path):
    rb = ReportBuilder(year=2024)
    # dividends, withholding, syep interest
    rb.set_dividends(
        [
            DividendRow(
                currency="USD",
                date=dt.date(2024, 1, 1),
                description="Test Div",
                amount=Decimal("100"),
            )
        ]
    )
    rb.set_withholding(
        [
            WithholdingRow(
                currency="USD",
                date=dt.date(2024, 1, 1),
                description="Test WHT",
                amount=Decimal("50"),
                code="WHT",
                type="",
                country="",
            )
        ]
    )
    rb.set_syep_interest(
        [
            SyepInterestRow(
                currency="USD",
                value_date=dt.date(2024, 1, 1),
                symbol="ABC",
                start_date=dt.date(2024, 1, 1),
                quantity=Decimal("-10"),
                collateral_amount=Decimal("1000"),
                market_rate_pct=Decimal("0.1"),
                customer_rate_pct=Decimal("0.05"),
                interest_paid=Decimal("1.23"),
                code="Po",
            )
        ]
    )
    fx = make_fx({("USD", "2024-01-01"): Decimal("0.9")})
    rb.convert_eur(fx)
    assert rb.dividends[0].amount_eur == Decimal("90.00")
    assert rb.withholding[0].amount_eur == Decimal("45.00")
    assert rb.syep_interest[0].interest_paid_eur == Decimal("1.11")

    # Write and verify SYEP sheet headers include EUR column
    out = tmp_path / "out.xlsx"
    sink = ExcelReportSink(out_path=out, locale="EN")
    sink.write(rb)

    wb = load_workbook(out)
    assert "SYEP Interest" in wb.sheetnames
    ws = wb["SYEP Interest"]
    headers = [c.value for c in ws[1]]
    assert "Interest Paid (EUR)" in headers


def test_per_symbol_summary_trade_and_eur(tmp_path):
    rb = ReportBuilder(year=2024)
    # Build one realized line in USD
    from capitangains.reporting.fifo_domain import SellMatchLeg

    legs = [
        SellMatchLeg(
            buy_date=dt.date(2024, 1, 1),
            qty=Decimal("10"),
            lot_qty_before=Decimal("10"),
            alloc_cost_ccy=Decimal("800"),
        )
    ]
    rl = RealizedLine(
        symbol="GOOGL",
        currency="USD",
        sell_date=dt.date(2024, 2, 1),
        sell_qty=Decimal("10"),
        sell_gross_ccy=Decimal("1000"),
        sell_comm_ccy=Decimal("0"),
        sell_net_ccy=Decimal("1000"),
        legs=legs,
        realized_pl_ccy=Decimal("200.00"),
    )
    rb.add_realized(rl)

    fx = make_fx(
        {
            ("USD", "2024-02-01"): Decimal("0.9"),
            ("USD", "2024-01-01"): Decimal("0.9"),
        }
    )
    rb.convert_eur(fx)

    out = tmp_path / "out2.xlsx"
    sink = ExcelReportSink(out_path=out, locale="EN")
    sink.write(rb)

    wb = load_workbook(out)
    ws = wb["Per Symbol Summary"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if row[0] == "GOOGL":
            assert row[1] == "USD"
            assert isinstance(row[2], (int, float)) and float(row[2]) == 200.0
            assert isinstance(row[3], (int, float)) and float(row[3]) == 1000.0
            assert isinstance(row[4], (int, float)) and float(row[4]) == 800.0
            assert isinstance(row[5], (int, float)) and float(row[5]) == 180.0
            assert isinstance(row[6], (int, float)) and float(row[6]) == 900.0
            assert isinstance(row[7], (int, float)) and float(row[7]) == 720.0
            break
    else:
        raise AssertionError("GOOGL row not found in Per Symbol Summary")
