from __future__ import annotations

import datetime as dt
import json
from abc import ABC, abstractmethod
from collections.abc import Callable, Iterable, Iterator, Sequence
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Generic, Protocol, TypeVar

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from capitangains.conv import EUR, Currency

from .extract import SyepInterestRow, WithholdingRow
from .fifo_domain import RealizedLine, TransferProtocol
from .i18n import NumberFormats, labels_for
from .money import quantize_money
from .quadro_8a import Quadro8ALine
from .report_builder import ReportBuilder


def _gap_status(rl: RealizedLine) -> str:
    """Human-readable basis provenance for a realized line's status column.

    Distinguishes a clean FIFO match (empty) from the two gap outcomes: a residual lot
    synthesized from IBKR's Basis versus an unmatched remainder left at zero cost.
    gap_fixed implies has_gap, so it is checked first.
    """
    if not rl.has_gap:
        return ""
    if rl.gap_fixed:
        return "synthesized from Basis"
    return "zero-cost gap"


_RowT = TypeVar("_RowT")


@dataclass(frozen=True)
class _Column(ABC, Generic[_RowT]):
    """One column of a uniform table sheet: a header-label key plus the two behaviors
    that must change together when the column's rendering changes -- the cell-value
    coercion and the number format. Bundling them in one polymorphic class is the SRP
    unit ("how to render one kind of cell") and lets the engine apply both without
    switching on a format kind.
    """

    header_key: str

    @abstractmethod
    def cell_value(self, row: _RowT) -> object: ...

    def number_format(self, formats: NumberFormats, row: _RowT) -> str | None:
        # None leaves the cell at openpyxl's default "General" format. Returning None
        # (never "") and assigning only when non-None keeps TEXT cells untouched, a
        # load-bearing byte-preservation detail.
        return None


@dataclass(frozen=True)
class _TextColumn(_Column[_RowT]):
    value: Callable[[_RowT], object]

    def cell_value(self, row: _RowT) -> object:
        return self.value(row)


@dataclass(frozen=True)
class _DateColumn(_Column[_RowT]):
    value: Callable[[_RowT], dt.date | None]

    def cell_value(self, row: _RowT) -> object:
        return self.value(row)

    def number_format(self, formats: NumberFormats, row: _RowT) -> str | None:
        # Applied unconditionally, including to a None date (e.g. absent SYEP dates).
        return formats.date


@dataclass(frozen=True)
class _QtyColumn(_Column[_RowT]):
    value: Callable[[_RowT], Decimal]

    def cell_value(self, row: _RowT) -> object:
        return float(self.value(row))

    def number_format(self, formats: NumberFormats, row: _RowT) -> str | None:
        return "0.########"


@dataclass(frozen=True)
class _PctColumn(_Column[_RowT]):
    value: Callable[[_RowT], Decimal]

    def cell_value(self, row: _RowT) -> object:
        return float(self.value(row))

    def number_format(self, formats: NumberFormats, row: _RowT) -> str | None:
        return "0.00####"


@dataclass(frozen=True)
class _MoneyColumn(_Column[_RowT]):
    value: Callable[[_RowT], Decimal | None]
    currency: Callable[[_RowT], Currency]

    def cell_value(self, row: _RowT) -> object:
        # The sink is the boundary where money becomes cents. Some sources are raw
        # 8-decimal allocation pieces (sum of leg.alloc_cost_ccy from round_cost_piece),
        # so quantize here rather than store sub-cent noise the money format would mask.
        v = self.value(row)
        return None if v is None else float(quantize_money(v))

    def number_format(self, formats: NumberFormats, row: _RowT) -> str | None:
        # The currency selector is a uniform callable (per-row Currency or the EUR
        # constant), so no isinstance branch is needed. money() takes the code string,
        # so unwrap the Currency here, at the sink boundary.
        return formats.money(str(self.currency(row)))


def _legs_json(rl: RealizedLine) -> str:
    """Serialize a realized line's matched buy lots for its JSON cell."""
    return json.dumps(
        [
            {
                "buy_date": (leg.buy_date.isoformat() if leg.buy_date else None),
                "qty": str(leg.qty),
                "alloc_cost_ccy": str(leg.alloc_cost_ccy),
            }
            for leg in rl.legs
        ]
    )


@dataclass(frozen=True)
class _AnexoJRow:
    """One acquisition-leg row of the Annex J breakdown (a realized line flattened over
    its legs) with the per-leg EUR profit/loss precomputed.
    """

    symbol: str
    currency: Currency
    buy_date: dt.date | None
    sell_date: dt.date
    qty: Decimal
    alloc_eur: Decimal | None
    proceeds_eur: Decimal | None
    pl_eur: Decimal | None
    transferred: bool
    synthetic: bool


def _anexo_j_rows(report: ReportBuilder) -> Iterator[_AnexoJRow]:
    for rl in report.realized_lines:
        for leg in rl.legs:
            alloc_eur = leg.alloc_cost_eur
            proceeds_eur = leg.proceeds_share_eur
            # P/L only when both EUR operands exist; keep the proceeds - alloc order.
            pl_eur = (
                quantize_money(proceeds_eur - alloc_eur)
                if alloc_eur is not None and proceeds_eur is not None
                else None
            )
            yield _AnexoJRow(
                symbol=rl.symbol,
                currency=rl.currency,
                buy_date=leg.buy_date,
                sell_date=rl.sell_date,
                qty=leg.qty,
                alloc_eur=alloc_eur,
                proceeds_eur=proceeds_eur,
                pl_eur=pl_eur,
                transferred=leg.transferred,
                synthetic=leg.synthetic,
            )


@dataclass(frozen=True)
class _PerSymbolRow:
    """One symbol's totals (single trade currency plus EUR) with a gap flag."""

    symbol: str
    currency: Currency
    pl_tcy: Decimal
    net_tcy: Decimal
    alloc_tcy: Decimal
    pl_eur: Decimal
    net_eur: Decimal
    alloc_eur: Decimal
    has_gap: bool


def _per_symbol_rows(report: ReportBuilder) -> Iterator[_PerSymbolRow]:
    # A symbol's totals silently fold in any gap-filled or synthesized line; flag the
    # symbol so the aggregate isn't read as a clean FIFO result. Built once per call.
    gap_symbols = {rl.symbol for rl in report.realized_lines if rl.has_gap}
    # Invariant: each symbol maps to exactly one trade currency (enforced at ingestion),
    # so the single by_currency entry is taken directly.
    for symbol, totals in sorted(report.symbol_totals.items()):
        ccy, ccy_totals = next(iter(totals.by_currency.items()))
        yield _PerSymbolRow(
            symbol=symbol,
            currency=ccy,
            pl_tcy=ccy_totals.realized,
            net_tcy=ccy_totals.proceeds,
            alloc_tcy=ccy_totals.alloc_cost,
            pl_eur=totals.eur.realized,
            net_eur=totals.eur.proceeds,
            alloc_eur=totals.eur.alloc_cost,
            has_gap=symbol in gap_symbols,
        )


@dataclass(frozen=True)
class _SheetSpec(Generic[_RowT]):
    """Declarative description of one uniform table sheet.

    A column is defined exactly once, with its header label, value extractor, and format
    behavior bundled in one _Column; and column order lives in the single ordered
    columns list, so the header, value, and format for a column cannot desynchronize.
    rows owns all per-sheet preparation (sorting, flattening legs over realized lines,
    per-symbol aggregation) so the engine stays a pure header+row+format loop.

    label_section is separate from sheet_key because they usually match but not always
    (the SYEP sheet titles under "syep_interest" while its columns live under
    "syep"). skip_if_empty is evaluated on the prepared rows; this equals a
    source-collection check only because every provider is a 1:1 sort/passthrough; a
    future filtering provider would need a source-level predicate instead.

    """

    sheet_key: str
    label_section: str
    columns: Sequence[_Column[_RowT]]
    rows: Callable[[ReportBuilder], Iterable[_RowT]]
    skip_if_empty: bool = False


_REALIZED_SPEC: _SheetSpec[RealizedLine] = _SheetSpec(
    sheet_key="realized",
    label_section="realized",
    rows=lambda report: report.realized_lines,
    columns=(
        _TextColumn[RealizedLine]("ticker", value=lambda rl: rl.symbol),
        _TextColumn[RealizedLine]("trade_currency", value=lambda rl: str(rl.currency)),
        _DateColumn[RealizedLine]("sell_date", value=lambda rl: rl.sell_date),
        _QtyColumn[RealizedLine]("qty_sold", value=lambda rl: rl.sell_qty),
        _MoneyColumn[RealizedLine](
            "gross_tcy",
            value=lambda rl: rl.sell_gross_ccy,
            currency=lambda rl: rl.currency,
        ),
        _MoneyColumn[RealizedLine](
            "fees_tcy",
            value=lambda rl: rl.sell_comm_ccy,
            currency=lambda rl: rl.currency,
        ),
        _MoneyColumn[RealizedLine](
            "net_tcy",
            value=lambda rl: rl.sell_net_ccy,
            currency=lambda rl: rl.currency,
        ),
        _MoneyColumn[RealizedLine](
            "alloc_tcy",
            value=lambda rl: sum((leg.alloc_cost_ccy for leg in rl.legs), Decimal("0")),
            currency=lambda rl: rl.currency,
        ),
        _MoneyColumn[RealizedLine](
            "pl_tcy",
            value=lambda rl: rl.realized_pl_ccy,
            currency=lambda rl: rl.currency,
        ),
        _MoneyColumn[RealizedLine](
            "gross_eur",
            value=lambda rl: rl.sell_gross_eur,
            currency=lambda _r: EUR,
        ),
        _MoneyColumn[RealizedLine](
            "fees_eur",
            value=lambda rl: rl.sell_comm_eur,
            currency=lambda _r: EUR,
        ),
        _MoneyColumn[RealizedLine](
            "net_eur",
            value=lambda rl: rl.sell_net_eur,
            currency=lambda _r: EUR,
        ),
        _MoneyColumn[RealizedLine](
            "alloc_eur",
            value=lambda rl: rl.alloc_cost_eur,
            currency=lambda _r: EUR,
        ),
        _MoneyColumn[RealizedLine](
            "pl_eur",
            value=lambda rl: rl.realized_pl_eur,
            currency=lambda _r: EUR,
        ),
        _TextColumn[RealizedLine]("legs_json", value=_legs_json),
        _TextColumn[RealizedLine]("gap_status", value=_gap_status),
    ),
)


_ANEXO_J_SPEC: _SheetSpec[_AnexoJRow] = _SheetSpec(
    sheet_key="anexo_j",
    label_section="anexo_j",
    rows=_anexo_j_rows,
    columns=(
        _TextColumn[_AnexoJRow]("ticker", value=lambda r: r.symbol),
        _TextColumn[_AnexoJRow]("trade_currency", value=lambda r: str(r.currency)),
        _DateColumn[_AnexoJRow]("buy_date", value=lambda r: r.buy_date),
        _DateColumn[_AnexoJRow]("sell_date", value=lambda r: r.sell_date),
        _QtyColumn[_AnexoJRow]("qty", value=lambda r: r.qty),
        _MoneyColumn[_AnexoJRow](
            "alloc_eur", value=lambda r: r.alloc_eur, currency=lambda _r: EUR
        ),
        _MoneyColumn[_AnexoJRow](
            "proceeds_eur", value=lambda r: r.proceeds_eur, currency=lambda _r: EUR
        ),
        _MoneyColumn[_AnexoJRow](
            "pl_eur", value=lambda r: r.pl_eur, currency=lambda _r: EUR
        ),
        _TextColumn[_AnexoJRow](
            "transferred", value=lambda r: "Yes" if r.transferred else ""
        ),
        _TextColumn[_AnexoJRow](
            "synthetic", value=lambda r: "Yes" if r.synthetic else ""
        ),
    ),
)


def _quadro_8a_rows(report: ReportBuilder) -> Iterable[Quadro8ALine]:
    return report.quadro_8a  # already sorted by (income code, kind, country)


def _quadro_8a_spec(kind_labels: dict[str, str]) -> _SheetSpec[Quadro8ALine]:
    """Build the Quadro 8A sheet spec for one locale.

    Unlike the other specs (module constants), this one is constructed per write because
    its Type column renders a localized row *value* (Dividend / Payment in Lieu /
    Interest) resolved from the active locale's labels, not just a localized header. The
    value lambda closes over those labels, keyed by IncomeKind.label_key.
    """
    return _SheetSpec(
        sheet_key="quadro_8a",
        label_section="quadro_8a",
        skip_if_empty=True,
        rows=_quadro_8a_rows,
        columns=(
            _TextColumn[Quadro8ALine]("income_code", value=lambda r: r.income_code),
            _TextColumn[Quadro8ALine](
                "kind", value=lambda r: kind_labels[r.kind.label_key]
            ),
            _TextColumn[Quadro8ALine]("country", value=lambda r: r.country),
            _MoneyColumn[Quadro8ALine](
                "gross_eur", value=lambda r: r.gross_eur, currency=lambda _r: EUR
            ),
            _MoneyColumn[Quadro8ALine](
                "tax_eur", value=lambda r: r.tax_eur, currency=lambda _r: EUR
            ),
        ),
    )


_PER_SYMBOL_SPEC: _SheetSpec[_PerSymbolRow] = _SheetSpec(
    sheet_key="per_symbol",
    label_section="per_symbol",
    rows=_per_symbol_rows,
    columns=(
        _TextColumn[_PerSymbolRow]("ticker", value=lambda r: r.symbol),
        _TextColumn[_PerSymbolRow]("trade_currency", value=lambda r: str(r.currency)),
        _MoneyColumn[_PerSymbolRow](
            "pl_tcy", value=lambda r: r.pl_tcy, currency=lambda r: r.currency
        ),
        _MoneyColumn[_PerSymbolRow](
            "net_tcy", value=lambda r: r.net_tcy, currency=lambda r: r.currency
        ),
        _MoneyColumn[_PerSymbolRow](
            "alloc_tcy", value=lambda r: r.alloc_tcy, currency=lambda r: r.currency
        ),
        _MoneyColumn[_PerSymbolRow](
            "pl_eur", value=lambda r: r.pl_eur, currency=lambda _r: EUR
        ),
        _MoneyColumn[_PerSymbolRow](
            "net_eur", value=lambda r: r.net_eur, currency=lambda _r: EUR
        ),
        _MoneyColumn[_PerSymbolRow](
            "alloc_eur", value=lambda r: r.alloc_eur, currency=lambda _r: EUR
        ),
        _TextColumn[_PerSymbolRow](
            "has_gap", value=lambda r: "Yes" if r.has_gap else ""
        ),
    ),
)


class _CashFlowRow(Protocol):
    """Shared shape of the dividend and account-interest rows: a dated, described cash
    flow in a trade currency with an optional EUR equivalent. Both render as the same
    five-column sheet, so a single _cash_flow_spec builds both.
    """

    date: dt.date
    currency: Currency
    description: str
    amount: Decimal
    amount_eur: Decimal | None


_CashFlowT = TypeVar("_CashFlowT", bound=_CashFlowRow)


def _cash_flow_spec(
    sheet_key: str, source: Callable[[ReportBuilder], Iterable[_CashFlowT]]
) -> _SheetSpec[_CashFlowT]:
    """Build the spec for a cash-flow sheet (dividends, account interest), sorted by
    description. The sheet title and column labels share sheet_key as their section.
    """
    return _SheetSpec(
        sheet_key=sheet_key,
        label_section=sheet_key,
        skip_if_empty=True,
        rows=lambda report: sorted(
            source(report), key=lambda row: row.description.lower()
        ),
        columns=(
            _DateColumn[_CashFlowT]("date", value=lambda r: r.date),
            _TextColumn[_CashFlowT]("currency", value=lambda r: str(r.currency)),
            _TextColumn[_CashFlowT]("desc", value=lambda r: r.description),
            _MoneyColumn[_CashFlowT](
                "amount", value=lambda r: r.amount, currency=lambda r: r.currency
            ),
            _MoneyColumn[_CashFlowT](
                "amount_eur", value=lambda r: r.amount_eur, currency=lambda _r: EUR
            ),
        ),
    )


_DIVIDENDS_SPEC = _cash_flow_spec("dividends", lambda report: report.dividends)
_INTEREST_SPEC = _cash_flow_spec("interest", lambda report: report.interest)


_SYEP_SPEC: _SheetSpec[SyepInterestRow] = _SheetSpec(
    sheet_key="syep_interest",
    label_section="syep",
    skip_if_empty=True,
    rows=lambda report: report.syep_interest,
    columns=(
        _DateColumn[SyepInterestRow]("date", value=lambda s: s.value_date),
        _TextColumn[SyepInterestRow]("currency", value=lambda s: str(s.currency)),
        _TextColumn[SyepInterestRow]("symbol", value=lambda s: s.symbol),
        _DateColumn[SyepInterestRow]("start_date", value=lambda s: s.start_date),
        _QtyColumn[SyepInterestRow]("quantity", value=lambda s: s.quantity),
        _MoneyColumn[SyepInterestRow](
            "collateral",
            value=lambda s: s.collateral_amount,
            currency=lambda s: s.currency,
        ),
        _PctColumn[SyepInterestRow]("market_rate", value=lambda s: s.market_rate_pct),
        _PctColumn[SyepInterestRow](
            "customer_rate", value=lambda s: s.customer_rate_pct
        ),
        _MoneyColumn[SyepInterestRow](
            "interest_paid",
            value=lambda s: s.interest_paid,
            currency=lambda s: s.currency,
        ),
        _MoneyColumn[SyepInterestRow](
            "interest_paid_eur",
            value=lambda s: s.interest_paid_eur,
            currency=lambda _s: EUR,
        ),
        _TextColumn[SyepInterestRow]("code", value=lambda s: s.code),
    ),
)


_WITHHOLDING_SPEC: _SheetSpec[WithholdingRow] = _SheetSpec(
    sheet_key="withholding",
    label_section="withholding",
    skip_if_empty=True,
    rows=lambda report: sorted(
        report.withholding,
        key=lambda row: (str(row.currency), row.description.lower()),
    ),
    columns=(
        _DateColumn[WithholdingRow]("date", value=lambda w: w.date),
        _TextColumn[WithholdingRow]("currency", value=lambda w: str(w.currency)),
        _TextColumn[WithholdingRow]("desc", value=lambda w: w.description),
        _TextColumn[WithholdingRow]("type", value=lambda w: w.type),
        _TextColumn[WithholdingRow]("country", value=lambda w: w.country),
        _MoneyColumn[WithholdingRow](
            "amount", value=lambda w: w.amount, currency=lambda w: w.currency
        ),
        _MoneyColumn[WithholdingRow](
            "amount_eur", value=lambda w: w.amount_eur, currency=lambda _w: EUR
        ),
    ),
)


_TRANSFERS_SPEC: _SheetSpec[TransferProtocol] = _SheetSpec(
    sheet_key="transfers",
    label_section="transfers",
    skip_if_empty=True,
    rows=lambda report: sorted(report.transfers, key=lambda t: (t.date, t.symbol)),
    columns=(
        _DateColumn[TransferProtocol]("date", value=lambda t: t.date),
        _TextColumn[TransferProtocol]("symbol", value=lambda t: t.symbol),
        _TextColumn[TransferProtocol]("direction", value=lambda t: t.direction),
        _QtyColumn[TransferProtocol]("quantity", value=lambda t: t.quantity),
        _TextColumn[TransferProtocol]("currency", value=lambda t: str(t.currency)),
        _MoneyColumn[TransferProtocol](
            "market_value",
            value=lambda t: t.market_value,
            currency=lambda t: t.currency,
        ),
        _TextColumn[TransferProtocol]("code", value=lambda t: t.code),
    ),
)


@dataclass(frozen=True)
class _SheetWriter:
    """One workbook under construction in a single locale.

    Bundles the openpyxl workbook with the locale-derived label view and number formats
    so the per-sheet writers do not re-thread that context on every call. Constructed
    fresh per ExcelReportSink.write; frozen because the bindings are fixed for one build
    even though the workbook they point at is mutated in place.
    """

    wb: Workbook
    labels: dict[str, dict[str, str]]
    formats: NumberFormats

    def write_summary(self, report: ReportBuilder) -> None:
        ws = self.wb.create_sheet(title=self.labels["sheet"]["summary"])
        total_eur = sum(
            (rl.realized_pl_eur or Decimal("0") for rl in report.realized_lines),
            Decimal("0"),
        )
        proceeds_total_eur = sum(
            (rl.sell_net_eur or Decimal("0") for rl in report.realized_lines),
            Decimal("0"),
        )
        alloc_total_eur = sum(
            (rl.alloc_cost_eur or Decimal("0") for rl in report.realized_lines),
            Decimal("0"),
        )

        totals_by_cur: dict[str, Decimal] = {}
        for rl in report.realized_lines:
            # Exclude EUR from by-currency totals to avoid duplicate label confusion
            if rl.currency.is_base:
                continue
            code = str(rl.currency)
            totals_by_cur[code] = (
                totals_by_cur.get(code, Decimal("0")) + rl.realized_pl_ccy
            )
        ws.append([self.labels["summary"]["metric"], self.labels["summary"]["amount"]])

        # Primary EUR totals
        ws.append([self.labels["summary"]["total_eur"], float(total_eur)])
        ws.cell(row=ws.max_row, column=2).number_format = self.formats.money(str(EUR))
        ws.append([self.labels["summary"]["proceeds_eur"], float(proceeds_total_eur)])
        ws.cell(row=ws.max_row, column=2).number_format = self.formats.money(str(EUR))
        ws.append([self.labels["summary"]["alloc_eur"], float(alloc_total_eur)])
        ws.cell(row=ws.max_row, column=2).number_format = self.formats.money(str(EUR))

        for cur, amt in sorted(totals_by_cur.items()):
            ws.append(
                [self.labels["summary"]["total_cur_tpl"].format(cur=cur), float(amt)]
            )
            ws.cell(row=ws.max_row, column=2).number_format = self.formats.money(cur)

    def write_table(self, spec: _SheetSpec[_RowT], report: ReportBuilder) -> None:
        """Write one uniform table sheet from its spec.

        Header, value, and format for column N all derive from the single ordered
        spec.columns[N], so they cannot desynchronize; the format is applied
        positionally over the same list that produced the value, and only when the
        column returns one (TEXT cells are left at openpyxl's default "General").
        """
        rows = list(spec.rows(report))
        if not rows and spec.skip_if_empty:
            return

        section = self.labels[spec.label_section]
        ws = self.wb.create_sheet(title=self.labels["sheet"][spec.sheet_key])
        ws.append([section[col.header_key] for col in spec.columns])

        for row in rows:
            ws.append([col.cell_value(row) for col in spec.columns])
            r = ws.max_row
            for idx, col in enumerate(spec.columns, start=1):
                fmt = col.number_format(self.formats, row)
                if fmt is not None:
                    ws.cell(row=r, column=idx).number_format = fmt

    def autosize_all(self) -> None:
        for ws in self.wb.worksheets:
            self._autosize(ws)

    def _autosize(
        self, sheet: Worksheet, max_width: int = 60, min_width: int = 10
    ) -> None:
        header_values = [cell.value for cell in sheet[1]] if sheet.max_row else []
        for col in range(1, sheet.max_column + 1):
            max_len = 0
            for row in range(1, sheet.max_row + 1):
                v = sheet.cell(row=row, column=col).value
                if v is None:
                    continue
                if hasattr(v, "strftime"):
                    s = (
                        v.strftime("%d/%m/%Y")
                        if self.formats.locale.upper() == "PT"
                        else v.strftime("%Y-%m-%d")
                    )
                else:
                    s = str(v)
                if len(s) > max_len:
                    max_len = len(s)

            header = header_values[col - 1] if col - 1 < len(header_values) else None
            if header:
                max_len = max(max_len, len(str(header)))
            width = min(max_width, max(min_width, max_len + 2))
            if header and "JSON" in str(header):
                width = min(width, 50)
            sheet.column_dimensions[get_column_letter(col)].width = width


@dataclass
class ExcelReportSink:
    out_path: Path
    locale: str = "PT"  # "PT" (default) or "EN"

    def write(self, report: ReportBuilder) -> Path:
        out_path = Path(self.out_path)
        wb = Workbook()

        ws_default = wb.active
        if ws_default is not None:
            wb.remove(ws_default)

        labels = labels_for(self.locale)
        writer = _SheetWriter(wb, labels, NumberFormats(self.locale))
        writer.write_summary(report)
        writer.write_table(_REALIZED_SPEC, report)
        writer.write_table(_ANEXO_J_SPEC, report)
        writer.write_table(_quadro_8a_spec(labels["quadro_8a"]), report)
        writer.write_table(_PER_SYMBOL_SPEC, report)
        writer.write_table(_DIVIDENDS_SPEC, report)
        writer.write_table(_INTEREST_SPEC, report)
        writer.write_table(_SYEP_SPEC, report)
        writer.write_table(_WITHHOLDING_SPEC, report)
        writer.write_table(_TRANSFERS_SPEC, report)
        writer.autosize_all()

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path
